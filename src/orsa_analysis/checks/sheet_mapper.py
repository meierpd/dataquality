"""Multi-language sheet name mapping for ORSA Excel files.

This module handles translation between German (reference), English, and French
sheet names in ORSA Excel files.
"""

from typing import Dict, Optional, List
from openpyxl.workbook.workbook import Workbook
import logging

logger = logging.getLogger(__name__)


SHEET_NAME_MAPPING: Dict[str, Dict[str, str]] = {
    "Mgmt. Summary": {
        "EN": "Mgmt. summary",
        "FR": "Mgmt. summary"
    },
    "Auswertung": {
        "EN": "General details",
        "FR": "Info. générales"
    },
    "Allgem. Angaben": {
        "EN": "Risks",
        "FR": "Risques"
    },
    "Risiken": {
        "EN": "Measures",
        "FR": "Mesures"
    },
    "Massnahmen": {
        "EN": "Scenarios",
        "FR": "Scénarios"
    },
    "Szenarien": {
        "EN": "Results_ISO-FINMA",
        "FR": "Résultats_OS-FINMA"
    },
    "Ergebnisse_AVO-FINMA": {
        "EN": "Results_ISO-FINMA",
        "FR": "Résultats_OS-FINMA"
    },
    "Ergebnisse_IFRS": {
        "EN": "Results_IFRS",
        "FR": "Résultats_IFRS"
    },
    "Qual. & langfr. Risiken": {
        "EN": "Qual. & long-term risks",
        "FR": "Risques qual. & à long terme"
    },
    "Schlussfolgerungen, Dokument.": {
        "EN": "Conclusions, documentation",
        "FR": "Conclusions, document."
    },
    "Drop-downs": {
        "EN": "Drop-downs",
        "FR": "Drop-Downs"
    }
}


class SheetNameMapper:
    """Maps German reference sheet names to actual sheet names in different languages."""

    def __init__(self, workbook: Workbook):
        """Initialize the mapper with a workbook.

        Args:
            workbook: The Excel workbook to analyze
        """
        self.workbook = workbook
        self.detected_language = self._detect_language()
        logger.info(f"Detected workbook language: {self.detected_language}")

    def _detect_language(self) -> str:
        """Detect the language of the workbook by checking sheet names.

        Returns:
            'DE' for German, 'EN' for English, 'FR' for French
        """
        sheet_names = set(self.workbook.sheetnames)
        
        # Count matches for each language
        de_matches = 0
        en_matches = 0
        fr_matches = 0
        
        for de_name, translations in SHEET_NAME_MAPPING.items():
            if de_name in sheet_names:
                de_matches += 1
            if translations["EN"] in sheet_names:
                en_matches += 1
            if translations["FR"] in sheet_names:
                fr_matches += 1
        
        # Determine language based on highest match count
        if de_matches >= en_matches and de_matches >= fr_matches:
            return "DE"
        elif en_matches >= fr_matches:
            return "EN"
        else:
            return "FR"

    def get_sheet_name(self, german_reference: str) -> Optional[str]:
        """Get the actual sheet name in the workbook's language.

        Args:
            german_reference: The German reference name

        Returns:
            The actual sheet name, or None if not found
        """
        if self.detected_language == "DE":
            # For German, return the reference name itself
            return german_reference if german_reference in self.workbook.sheetnames else None
        
        # For other languages, look up in mapping
        if german_reference not in SHEET_NAME_MAPPING:
            logger.warning(f"German reference '{german_reference}' not found in mapping")
            return None
        
        translated_name = SHEET_NAME_MAPPING[german_reference][self.detected_language]
        
        # Verify the translated name exists in the workbook
        if translated_name not in self.workbook.sheetnames:
            logger.warning(
                f"Translated sheet name '{translated_name}' ({self.detected_language}) "
                f"not found in workbook"
            )
            return None
        
        return translated_name

    def get_sheet(self, german_reference: str):
        """Get the worksheet object by German reference name.

        Args:
            german_reference: The German reference name

        Returns:
            The worksheet object, or None if not found
        """
        sheet_name = self.get_sheet_name(german_reference)
        if sheet_name is None:
            return None
        
        try:
            return self.workbook[sheet_name]
        except KeyError:
            logger.error(f"Sheet '{sheet_name}' not found in workbook")
            return None

    def has_sheet(self, german_reference: str) -> bool:
        """Check if a sheet exists by German reference name.

        Args:
            german_reference: The German reference name

        Returns:
            True if the sheet exists, False otherwise
        """
        return self.get_sheet_name(german_reference) is not None

    def get_all_mapped_sheets(self) -> Dict[str, str]:
        """Get all mapped sheets in the workbook.

        Returns:
            Dictionary mapping German reference names to actual sheet names
        """
        result = {}
        for german_name in SHEET_NAME_MAPPING.keys():
            actual_name = self.get_sheet_name(german_name)
            if actual_name:
                result[german_name] = actual_name
        return result
