"""Quality check rules and registry.

Each check function receives a Workbook object and returns a tuple:
(outcome_bool, outcome_numeric, description)
"""

import logging
import re
from datetime import date, datetime
from typing import Callable, Optional, Tuple

from openpyxl.workbook.workbook import Workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from .sheet_mapper import SheetNameMapper

logger = logging.getLogger(__name__)

CheckFunction = Callable[[Workbook], Tuple[bool, str, str]]

STR_GUT = "gut"
STR_UNGENUEGEND = "ungenügend"

### Allgemeine Angaben

def check_responsible_person(wb: Workbook) -> Tuple[bool, str, str]:
    """
    Get the responsible person from the workbook, to fill it into the Auswertungs file.
    This is not a real check, but needed to get the value into the report.
    """
    try:
        mapper = SheetNameMapper(wb)
        sheet = mapper.get_sheet("Allgem. Angaben")

        if sheet is None:
            return False, "NA", "Das Tabellenblatt 'Allgem. Angaben' wurde in der Arbeitsmappe nicht gefunden"

        value = sheet["C8"].value
        value = "" if value is None else str(value)

        return True, value, f"Verantwortliche Person aus Zelle C8: {value}" if value else "Keine verantwortliche Person in Zelle C8 angegeben"

    except Exception as e:
        logger.error(f"Error in check_responsible_person: {e}")
        return False, "NA", f"Prüfung fehlgeschlagen mit Fehler: {str(e)}"


def _to_date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, str):
        return datetime.strptime(v.strip(), "%d.%m.%Y").date()
    return None


def _months_diff(start: date, end: date) -> int:
    months = (end.year - start.year) * 12 + (end.month - start.month)
    if end.day < start.day:
        months -= 1
    return months


def check_data_recency_geschaeftsplanung(wb: Workbook) -> Tuple[bool, str, str]:
    """
    Check that the data which the orsa is based on is recent enough.
    """
    try:
        mapper = SheetNameMapper(wb)
        sheet = mapper.get_sheet("Allgem. Angaben")

        if sheet is None:
            return (
                False,
                STR_UNGENUEGEND,
                "Das Tabellenblatt 'Allgem. Angaben' wurde in der Arbeitsmappe nicht gefunden",
            )

        approved = _to_date(sheet["C17"].value)
        snapshot = _to_date(sheet["C14"].value)

        if approved is None:
            return (
                False,
                STR_UNGENUEGEND,
                "Genehmigungsdatum durch Verwaltungsrat (C17) fehlt oder ist ungültig",
            )
        if snapshot is None:
            return False, STR_UNGENUEGEND, "Stichtagsdatum der Geschäftsplanung (C14) fehlt oder ist ungültig"
        if approved < snapshot:
            return False, STR_UNGENUEGEND, "Genehmigungsdatum liegt vor dem Stichtagsdatum (unlogisch)"

        months = _months_diff(snapshot, approved)
        ok = months <= 6

        if ok:
            return True, STR_GUT, f"Aktualität der Daten für Geschäftsplanung ist ausreichend: {months} Monate zwischen Stichtag und Genehmigung (≤6 Monate erforderlich)"
        else:
            return False, STR_UNGENUEGEND, f"Aktualität der Daten für Geschäftsplanung ist ungenügend: {months} Monate zwischen Stichtag und Genehmigung (max. 6 Monate erlaubt)"

    except Exception as e:
        logger.error(f"Error in check_data_recency_geschaeftsplanung: {e}")
        return False, STR_UNGENUEGEND, f"Prüfung fehlgeschlagen mit Fehler: {str(e)}"


def check_data_recency_risikoidentifikation(wb: Workbook) -> Tuple[bool, str, str]:
    """
    Check that the data which the orsa is based on is recent enough.
    """
    try:
        mapper = SheetNameMapper(wb)
        sheet = mapper.get_sheet("Allgem. Angaben")

        if sheet is None:
            return (
                False,
                STR_UNGENUEGEND,
                "Das Tabellenblatt 'Allgem. Angaben' wurde in der Arbeitsmappe nicht gefunden",
            )

        approved = _to_date(sheet["C17"].value)
        snapshot = _to_date(sheet["C15"].value)

        if approved is None:
            return (
                False,
                STR_UNGENUEGEND,
                "Genehmigungsdatum durch Verwaltungsrat (C17) fehlt oder ist ungültig",
            )
        if snapshot is None:
            return False, STR_UNGENUEGEND, "Stichtagsdatum der Risikoidentifikation (C15) fehlt oder ist ungültig"
        if approved < snapshot:
            return False, STR_UNGENUEGEND, "Genehmigungsdatum liegt vor dem Stichtagsdatum (unlogisch)"

        months = _months_diff(snapshot, approved)
        ok = months <= 6

        if ok:
            return True, STR_GUT, f"Aktualität der Daten für Risikoidentifikation ist ausreichend: {months} Monate zwischen Stichtag und Genehmigung (≤6 Monate erforderlich)"
        else:
            return False, STR_UNGENUEGEND, f"Aktualität der Daten für Risikoidentifikation ist ungenügend: {months} Monate zwischen Stichtag und Genehmigung (max. 6 Monate erlaubt)"

    except Exception as e:
        logger.error(f"Error in check_data_recency_risikoidentifikation: {e}")
        return False, STR_UNGENUEGEND, f"Prüfung fehlgeschlagen mit Fehler: {str(e)}"


def check_data_recency_szenarien(wb: Workbook) -> Tuple[bool, str, str]:
    """
    Check that the data which the orsa is based on is recent enough.
    """
    try:
        mapper = SheetNameMapper(wb)
        sheet = mapper.get_sheet("Allgem. Angaben")

        if sheet is None:
            return (
                False,
                STR_UNGENUEGEND,
                "Das Tabellenblatt 'Allgem. Angaben' wurde in der Arbeitsmappe nicht gefunden",
            )

        approved = _to_date(sheet["C17"].value)
        snapshot = _to_date(sheet["C16"].value)

        if approved is None:
            return (
                False,
                STR_UNGENUEGEND,
                "Genehmigungsdatum durch Verwaltungsrat (C17) fehlt oder ist ungültig",
            )
        if snapshot is None:
            return False, STR_UNGENUEGEND, "Stichtagsdatum der Szenarien (C16) fehlt oder ist ungültig"
        if approved < snapshot:
            return False, STR_UNGENUEGEND, "Genehmigungsdatum liegt vor dem Stichtagsdatum (unlogisch)"

        months = _months_diff(snapshot, approved)
        ok = months <= 6

        if ok:
            return True, STR_GUT, f"Aktualität der Daten für Szenarien ist ausreichend: {months} Monate zwischen Stichtag und Genehmigung (≤6 Monate erforderlich)"
        else:
            return False, STR_UNGENUEGEND, f"Aktualität der Daten für Szenarien ist ungenügend: {months} Monate zwischen Stichtag und Genehmigung (max. 6 Monate erlaubt)"

    except Exception as e:
        logger.error(f"Error in check_data_recency_szenarien: {e}")
        return False, STR_UNGENUEGEND, f"Prüfung fehlgeschlagen mit Fehler: {str(e)}"


def check_board_approved_orsa(wb: Workbook) -> Tuple[bool, str, str]:
    # Check if this is a Zweigniederlassungs version
    if _is_zweigniederlassungs_version(wb):
        return False, "kein Rating", "Kein Rating da es sich um eine Zweigniederlassung handelt"
    
    mapper = SheetNameMapper(wb)
    sheet = mapper.get_sheet("Allgem. Angaben")
    
    if sheet is None:
        return False, STR_UNGENUEGEND, "Das Tabellenblatt 'Allgem. Angaben' wurde in der Arbeitsmappe nicht gefunden"

    value_who_approved = sheet["C18"].value or ""
    is_approved_through_board = value_who_approved.startswith("(1)") or value_who_approved.startswith("(2)")


    if is_approved_through_board:
        return True, STR_GUT, f"ORSA wurde durch den Verwaltungsrat genehmigt: {value_who_approved}"
    else:
        return False, STR_UNGENUEGEND, f"ORSA wurde nicht durch den Verwaltungsrat genehmigt. Gefundener Wert in C18: '{value_who_approved}' (erwartet: '(1) gesamter VR' oder '(2) ein VR-Ausschuss')"

## Risiken

def check_risikobeurteilung_method(wb: Workbook) -> Tuple[bool, str, str]:
    mapper = SheetNameMapper(wb)
    sheet = mapper.get_sheet("Risiken")
    
    if sheet is None:
        return False, "Prüfen", "Das Tabellenblatt 'Risiken' wurde in der Arbeitsmappe nicht gefunden"

    value_method = sheet["E7"].value or ""

    if value_method.startswith("(1)"):
        result_str = "gut"
        desc = f"Risikobeurteilungsmethode ist gut: {value_method}"
    elif value_method.startswith("(2)"):
        result_str = "mangelhaft"
        desc = f"Risikobeurteilungsmethode ist mangelhaft: {value_method}"
    elif value_method.startswith("(3)"):
        result_str = "ungenügend"
        desc = f"Risikobeurteilungsmethode ist ungenügend: {value_method}"
    elif value_method.startswith("(4)"):
        result_str = "kein Rating"
        desc = f"Kein Rating für Risikobeurteilungsmethode: {value_method}"
    else:
        result_str = ""
        desc = f"Ungültiger oder fehlender Wert für Risikobeurteilungsmethode in Zelle E7: '{value_method}'"

    return result_str == "gut", result_str, desc


def check_risk_criteria_sufficient(wb: Workbook) -> Tuple[bool, str, str]:
    mapper = SheetNameMapper(wb)
    sheet = mapper.get_sheet("Risiken")
    
    if sheet is None:
        return False, "mangelhaft", "Das Tabellenblatt 'Risiken' wurde in der Arbeitsmappe nicht gefunden"

    # Check if this is a Zweigniederlassungs version
    is_zweigniederlassung = _is_zweigniederlassungs_version(wb)
    
    # Collect all criteria found
    all_criteria = {"(1)", "(2)", "(3)", "(4)", "(5)"}
    found = set()

    for row in range(10, 17):
        value = sheet[f"E{row}"].value or ""
        for criterion in all_criteria:
            if value.startswith(criterion):
                found.add(criterion)

    if is_zweigniederlassung:
        # For Zweigniederlassung:
        # 1, 2 are present -> genügend
        # 1, 2 + more is present -> gut
        # 1 or 2 not present -> mangelhaft
        has_1 = "(1)" in found
        has_2 = "(2)" in found
        
        if has_1 and has_2:
            # Both 1 and 2 are present
            if len(found) > 2:
                return True, "gut", f"Risikokriterien sind gut (Zweigniederlassung): (1) und (2) vorhanden, plus weitere Kriterien. Gefunden: {', '.join(sorted(found))}"
            else:
                return True, "genügend", f"Risikokriterien sind genügend (Zweigniederlassung): (1) und (2) vorhanden. Gefunden: {', '.join(sorted(found))}"
        else:
            missing = []
            if not has_1:
                missing.append("(1)")
            if not has_2:
                missing.append("(2)")
            return False, "mangelhaft", f"Risikokriterien sind mangelhaft (Zweigniederlassung): {', '.join(missing)} fehlt/fehlen. Gefunden: {', '.join(sorted(found)) if found else 'keine'}"
    else:
        # For Sitzgesellschaft:
        # (1 oder 2 is present), 3 4 and 5 is present -> genügend
        # (1 oder 2 is present), 3 4 and 5 is present + more numbers present -> gut
        # otherwise mangelhaft
        has_1_or_2 = "(1)" in found or "(2)" in found
        has_3 = "(3)" in found
        has_4 = "(4)" in found
        has_5 = "(5)" in found
        
        if has_1_or_2 and has_3 and has_4 and has_5:
            # Has (1 or 2) and 3, 4, 5
            if len(found) == 4:
                # Exactly (1 or 2), 3, 4, 5
                return True, "genügend", f"Risikokriterien sind genügend (Sitzgesellschaft): (1) oder (2) sowie (3), (4), (5) vorhanden. Gefunden: {', '.join(sorted(found))}"
            else:
                # More than 4 criteria
                return True, "gut", f"Risikokriterien sind gut (Sitzgesellschaft): (1) oder (2) sowie (3), (4), (5) vorhanden, plus weitere Kriterien. Gefunden: {', '.join(sorted(found))}"
        else:
            missing = []
            if not has_1_or_2:
                missing.append("(1) oder (2)")
            if not has_3:
                missing.append("(3)")
            if not has_4:
                missing.append("(4)")
            if not has_5:
                missing.append("(5)")
            return False, "mangelhaft", f"Risikokriterien sind mangelhaft (Sitzgesellschaft): {', '.join(missing)} fehlt/fehlen. Gefunden: {', '.join(sorted(found)) if found else 'keine'}"


def _count_risk(wb: Workbook, prefix: str) -> str:
    mapper = SheetNameMapper(wb)
    sheet = mapper.get_sheet("Risiken")
    
    if sheet is None:
        return "0"

    count = 0
    for row in range(22, 52):
        value = sheet[f"E{row}"].value or ""
        if value.startswith(prefix):
            count += 1

    return str(count)


def check_finanzmarktrisiko_count(wb: Workbook) -> Tuple[bool, str, str]:
    count_str = _count_risk(wb, "(1)")
    return True, count_str, f"Anzahl identifizierter Finanzmarktrisiken: {count_str}"


def check_versicherungsrisiko_count(wb: Workbook) -> Tuple[bool, str, str]:
    count_str = _count_risk(wb, "(2)")
    return True, count_str, f"Anzahl identifizierter Versicherungsrisiken: {count_str}"


def check_kreditrisiko_count(wb: Workbook) -> Tuple[bool, str, str]:
    count_str = _count_risk(wb, "(3)")
    return True, count_str, f"Anzahl identifizierter Kreditrisiken: {count_str}"


def check_liquiditaetsrisiko_count(wb: Workbook) -> Tuple[bool, str, str]:
    count_str = _count_risk(wb, "(4)")
    return True, count_str, f"Anzahl identifizierter Liquiditätsrisiken: {count_str}"


def check_operationelles_risiko_count(wb: Workbook) -> Tuple[bool, str, str]:
    count_str = _count_risk(wb, "(5)")
    return True, count_str, f"Anzahl identifizierter operationeller Risiken: {count_str}"


def check_strategisches_umfeld_risiko_count(wb: Workbook) -> Tuple[bool, str, str]:
    count_str = _count_risk(wb, "(6)")
    return True, count_str, f"Anzahl identifizierter strategischer/Umfeld-Risiken: {count_str}"


def check_anderes_risiko_count(wb: Workbook) -> Tuple[bool, str, str]:
    count_str = _count_risk(wb, "(7)")
    return True, count_str, f"Anzahl identifizierter anderer Risiken: {count_str}"

## Massnahmen

def check_count_number_mitigating_measures(wb: Workbook) -> Tuple[bool, str, str]:
    mapper = SheetNameMapper(wb)
    sheet = mapper.get_sheet("Massnahmen")
    
    if sheet is None:
        return True, "0", "Anzahl risikobegrenzender Massnahmen: 0 (Tabellenblatt 'Massnahmen' nicht gefunden)"

    # Count entries with (1)-(5) in column E (rows 9-38)
    count = 0
    for row in range(9, 39):
        value = str(sheet[f"E{row}"].value or "")
        if any(value.startswith(f"({i})") for i in range(1, 6)):
            count += 1

    count_str = str(count)
    return True, count_str, f"Anzahl risikobegrenzender Massnahmen: {count_str}"


def check_count_number_potential_mitigating_measures(
    wb: Workbook,
) -> Tuple[bool, str, str]:
    mapper = SheetNameMapper(wb)
    sheet = mapper.get_sheet("Massnahmen")
    
    if sheet is None:
        return True, "0", "Anzahl potenzieller risikobegrenzender Massnahmen: 0 (Tabellenblatt 'Massnahmen' nicht gefunden)"

    # Count entries with (1)-(5) in column G (rows 9-38)
    count = 0
    for row in range(9, 39):
        value = str(sheet[f"G{row}"].value or "")
        if any(value.startswith(f"({i})") for i in range(1, 6)):
            count += 1

    count_str = str(count)
    return True, count_str, f"Anzahl potenzieller risikobegrenzender Massnahmen: {count_str}"


def check_count_other_measures(wb: Workbook) -> Tuple[bool, str, str]:
    mapper = SheetNameMapper(wb)
    sheet = mapper.get_sheet("Massnahmen")
    
    if sheet is None:
        return True, "0", "Anzahl sonstiger Massnahmen: 0 (Tabellenblatt 'Massnahmen' nicht gefunden)"

    count = 0
    for row in range(44, 54):
        if (sheet[f"C{row}"].value or "") != "":
            count += 1

    count_str = str(count)
    return True, count_str, f"Anzahl sonstiger Massnahmen: {count_str}"


def check_count_potential_other_measures(wb: Workbook) -> Tuple[bool, str, str]:
    mapper = SheetNameMapper(wb)
    sheet = mapper.get_sheet("Massnahmen")
    
    if sheet is None:
        return True, "0", "Anzahl potenzieller sonstiger Massnahmen: 0 (Tabellenblatt 'Massnahmen' nicht gefunden)"

    count = 0
    for row in range(57, 67):
        if (sheet[f"C{row}"].value or "") != "":
            count += 1

    count_str = str(count)
    return True, count_str, f"Anzahl potenzieller sonstiger Massnahmen: {count_str}"

def check_risks_are_all_mitigated(wb: Workbook) -> Tuple[bool, str, str]:
    mapper = SheetNameMapper(wb)
    sheet_risiken = mapper.get_sheet("Risiken")
    sheet_massnahmen = mapper.get_sheet("Massnahmen")

    if sheet_risiken is None or sheet_massnahmen is None:
        return False, "NOK", "Tabellenblatt 'Risiken' oder 'Massnahmen' nicht gefunden"

    required_ids = set()
    for row in range(22, 52):
        if (sheet_risiken[f"C{row}"].value or "") != "":
            rid = sheet_risiken[f"B{row}"].value
            if rid is not None:
                required_ids.add(str(rid))

    mitigated_ids = set()
    # regex extracts a leading number like "1" from strings starting with "(1) ..."
    risk_id_re = re.compile(r"^\s*\(?\s*(\d+)\s*\)?")

    for row in range(9, 39):
        value = sheet_massnahmen[f"C{row}"].value or ""
        right_cell = sheet_massnahmen[f"D{row}"].value
        if right_cell is None or str(right_cell).strip() == "":
            continue

        m = risk_id_re.match(str(value))
        if m:
            mitigated_ids.add(m.group(1))

    missing = sorted(required_ids - mitigated_ids, key=int)

    if not missing:
        return True, "OK", f"Alle {len(required_ids)} identifizierten Risiken haben mindestens eine zugeordnete Massnahme"

    not_mitigated = ", ".join(missing)
    return False, "NOK", f"Risiken ohne zugeordnete Massnahmen (IDs): {not_mitigated}"


def check_any_nonmitigating_measures(wb: Workbook) -> Tuple[bool, str, str]:
    mapper = SheetNameMapper(wb)
    sheet = mapper.get_sheet("Massnahmen")
    
    if sheet is None:
        return True, "OK", "Tabellenblatt 'Massnahmen' nicht gefunden - keine nicht risikobegrenzenden Massnahmen vorhanden"

    count = 0
    for row in range(9, 39):
        value = sheet[f"E{row}"].value or ""
        if str(value).startswith("(5)"):
            count += 1

    count_str = str(count)
    is_ok = count == 0
    outcome_str = "OK" if is_ok else "NOK"

    if is_ok:
        return True, outcome_str, "Alle Massnahmen sind risikobegrenzend (keine mit Kategorie (5) 'Nicht risikobegrenzende Massnahme' gefunden)"
    else:
        return False, outcome_str, f"Anzahl nicht risikobegrenzender Massnahmen (Kategorie (5)): {count_str}. Diese Massnahmen sollten überprüft werden."


#### Szenarien
def _scenario_type_cells() -> list[str]:
    return [
        "C10",
        "C34",
        "C58",
        "C82",
        "C106",
        "C130",
        "C154",
        "C178",
        "C202",
        "C226",
        "C250",
        "C274",
        "C298",
        "C322",
        "C346",
    ]


def check_count_all_scenarios(wb: Workbook) -> Tuple[bool, str, str]:
    mapper = SheetNameMapper(wb)
    sheet = mapper.get_sheet("Szenarien")
    
    if sheet is None:
        return True, "0", "Anzahl aller definierten Szenarien: 0 (Tabellenblatt 'Szenarien' nicht gefunden)"

    count = 0
    for addr in _scenario_type_cells():
        if (sheet[addr].value or "") != "":
            count += 1

    count_str = str(count)
    return True, count_str, f"Anzahl aller definierten Szenarien: {count_str}"


def check_count_advers_scenarios(wb: Workbook) -> Tuple[bool, str, str]:
    mapper = SheetNameMapper(wb)
    sheet = mapper.get_sheet("Szenarien")
    
    if sheet is None:
        return True, "0", "Anzahl adverser Szenarien (Typ (1)): 0 (Tabellenblatt 'Szenarien' nicht gefunden)"

    count = 0
    for addr in _scenario_type_cells():
        value = str(sheet[addr].value or "")
        if value.startswith("(1)"):
            count += 1

    count_str = str(count)
    return True, count_str, f"Anzahl adverser Szenarien (Typ (1)): {count_str}"


def check_count_existenzbedrohend_scenarios(wb: Workbook) -> Tuple[bool, str, str]:
    mapper = SheetNameMapper(wb)
    sheet = mapper.get_sheet("Szenarien")
    
    if sheet is None:
        return True, "0", "Anzahl existenzbedrohender Szenarien (Typ (2)): 0 (Tabellenblatt 'Szenarien' nicht gefunden)"

    count = 0
    for addr in _scenario_type_cells():
        value = str(sheet[addr].value or "")
        if value.startswith("(2)"):
            count += 1

    count_str = str(count)
    return True, count_str, f"Anzahl existenzbedrohender Szenarien (Typ (2)): {count_str}"


def check_count_other_scenarios(wb: Workbook) -> Tuple[bool, str, str]:
    mapper = SheetNameMapper(wb)
    sheet = mapper.get_sheet("Szenarien")
    
    if sheet is None:
        return True, "0", "Anzahl sonstiger Szenarien (Typ (3)): 0 (Tabellenblatt 'Szenarien' nicht gefunden)"

    count = 0
    for addr in _scenario_type_cells():
        value = str(sheet[addr].value or "")
        if value.startswith("(3)"):
            count += 1

    count_str = str(count)
    return True, count_str, f"Anzahl sonstiger Szenarien (Typ (3)): {count_str}"

def check_every_scenario_has_event(wb: Workbook) -> Tuple[bool, str, str]:
    mapper = SheetNameMapper(wb)
    sheet = mapper.get_sheet("Szenarien")
    
    if sheet is None:
        return True, "OK", "Tabellenblatt 'Szenarien' nicht gefunden"

    ok = True
    for i, type_addr in enumerate(_scenario_type_cells()):
        scenario_type = sheet[type_addr].value or ""
        if scenario_type == "":
            continue

        start_row = 14 + i * 24
        has_event = False
        for r in range(start_row, start_row + 5):  # C14-C18, then +24 each scenario
            if (sheet[f"C{r}"].value or "") != "":
                has_event = True
                break

        if not has_event:
            ok = False
            break

    if ok:
        return True, "OK", "Jedes definierte Szenario hat mindestens ein zugeordnetes Ereignis"
    return False, "NOK", "Mindestens ein Szenario hat kein zugeordnetes Ereignis."

def check_count_scenarios_only_one_event(wb: Workbook) -> Tuple[bool, str, str]:
    mapper = SheetNameMapper(wb)
    sheet = mapper.get_sheet("Szenarien")
    
    if sheet is None:
        return True, "0", "Anzahl Szenarien mit genau einem Ereignis: 0 (Tabellenblatt 'Szenarien' nicht gefunden)"

    count = 0
    for i, type_addr in enumerate(_scenario_type_cells()):
        scenario_type = sheet[type_addr].value or ""
        if scenario_type == "":
            continue

        start_row = 14 + i * 24
        event_count = 0
        for r in range(start_row, start_row + 5):
            if (sheet[f"C{r}"].value or "") != "":
                event_count += 1

        if event_count == 1:
            count += 1

    count_str = str(count)
    return True, count_str, f"Anzahl Szenarien mit genau einem Ereignis: {count_str}"

def check_count_scenarios_multiple_events(wb: Workbook) -> Tuple[bool, str, str]:
    mapper = SheetNameMapper(wb)
    sheet = mapper.get_sheet("Szenarien")
    
    if sheet is None:
        return True, "0", "Anzahl Szenarien mit mehreren Ereignissen: 0 (Tabellenblatt 'Szenarien' nicht gefunden)"

    count = 0
    for i, type_addr in enumerate(_scenario_type_cells()):
        scenario_type = sheet[type_addr].value or ""
        if scenario_type == "":
            continue

        start_row = 14 + i * 24
        event_count = 0
        for r in range(start_row, start_row + 5):
            if (sheet[f"C{r}"].value or "") != "":
                event_count += 1

        if event_count > 1:
            count += 1

    count_str = str(count)
    return True, count_str, f"Anzahl Szenarien mit mehreren Ereignissen: {count_str}"

def check_every_scenario_has_risk(wb: Workbook) -> Tuple[bool, str, str]:
    mapper = SheetNameMapper(wb)
    sheet = mapper.get_sheet("Szenarien")
    
    if sheet is None:
        return True, "OK", "Tabellenblatt 'Szenarien' nicht gefunden"

    ok = True
    for i, type_addr in enumerate(_scenario_type_cells()):
        scenario_type = sheet[type_addr].value or ""
        if scenario_type == "":
            continue

        start_row = 20 + i * 24  # C20-C29, then +24 each scenario
        has_risk = False
        for r in range(start_row, start_row + 10):
            if (sheet[f"C{r}"].value or "") != "":
                has_risk = True
                break

        if not has_risk:
            ok = False
            break

    if ok:
        return True, "OK", "Jedes definierte Szenario hat mindestens ein zugeordnetes Risiko"
    return False, "NOK", "Mindestens ein Szenario hat kein zugeordnetes Risiko."

def check_count_scenarios_only_one_risk(wb: Workbook) -> Tuple[bool, str, str]:
    mapper = SheetNameMapper(wb)
    sheet = mapper.get_sheet("Szenarien")
    
    if sheet is None:
        return True, "0", "Anzahl Szenarien mit genau einem Risiko: 0 (Tabellenblatt 'Szenarien' nicht gefunden)"

    count = 0
    for i, type_addr in enumerate(_scenario_type_cells()):
        scenario_type = sheet[type_addr].value or ""
        if scenario_type == "":
            continue

        start_row = 20 + i * 24
        risk_count = 0
        for r in range(start_row, start_row + 10):
            if (sheet[f"C{r}"].value or "") != "":
                risk_count += 1

        if risk_count == 1:
            count += 1

    count_str = str(count)
    return True, count_str, f"Anzahl Szenarien mit genau einem Risiko: {count_str}"

def check_count_scenrios_multiple_risks(wb: Workbook) -> Tuple[bool, str, str]:
    mapper = SheetNameMapper(wb)
    sheet = mapper.get_sheet("Szenarien")
    
    if sheet is None:
        return True, "0", "Anzahl Szenarien mit mehreren Risiken: 0 (Tabellenblatt 'Szenarien' nicht gefunden)"

    count = 0
    for i, type_addr in enumerate(_scenario_type_cells()):
        scenario_type = sheet[type_addr].value or ""
        if scenario_type == "":
            continue

        start_row = 20 + i * 24
        risk_count = 0
        for r in range(start_row, start_row + 10):
            if (sheet[f"C{r}"].value or "") != "":
                risk_count += 1

        if risk_count > 1:
            count += 1

    count_str = str(count)
    return True, count_str, f"Anzahl Szenarien mit mehreren Risiken: {count_str}"

### Resultate AVO-FINMA / IFRS

def _is_any_filled(sheet, cells: list[str]) -> bool:
    if sheet is None:
        return False
    for addr in cells:
        v = sheet[addr].value
        if v is not None and str(v).strip() != "":
            return True
    return False


def _is_zweigniederlassungs_version(wb: Workbook) -> bool:
    """Detect if this is a Zweigniederlassungs version of the Excel file.
    
    The zweigniederlassungs_version has the sheet 'Ergebnisse' (and language equivalents),
    whereas the standard version has 'Ergebnisse_IFRS' and 'Ergebnisse_AVO-FINMA'.
    
    Args:
        wb: Workbook to check
        
    Returns:
        True if this is a Zweigniederlassungs version, False otherwise
    """
    sheet_names = set(wb.sheetnames)
    
    # Check for 'Ergebnisse' in all supported languages
    zweigniederlassungs_sheets = {
        "Ergebnisse",  # German
        "Results",      # English
        "Résultats"    # French
    }
    
    return any(name in sheet_names for name in zweigniederlassungs_sheets)
  
def _is_avo_finma_sheet(sheet_title: str) -> bool:
    """Check if a sheet title corresponds to the AVO-FINMA results sheet.
    
    Handles all supported languages (German, English, French).
    
    Args:
        sheet_title: The title of the sheet to check
        
    Returns:
        True if this is an AVO-FINMA results sheet, False otherwise
    """
    avo_finma_sheets = {
        "Ergebnisse_AVO-FINMA",  # German
        "Results_ISO-FINMA",      # English
        "Résultats_OS-FINMA"     # French
    }
    return sheet_title in avo_finma_sheets

def _get_filled_results_sheet(wb: Workbook) -> Tuple[bool, str, str, object]:
    mapper = SheetNameMapper(wb)
    
    # Check if this is a Zweigniederlassungs version
    if _is_zweigniederlassungs_version(wb):
        sheet_ergebnisse = mapper.get_sheet("Ergebnisse")
        if sheet_ergebnisse is None:
            msg = "Error: Ergebnisse sheet not found in Zweigniederlassungs version"
            return False, msg, msg, None
        return True, "OK", "OK", sheet_ergebnisse
    
    # Standard version logic
    sheet_avo = mapper.get_sheet("Ergebnisse_AVO-FINMA")
    sheet_ifrs = mapper.get_sheet("Ergebnisse_IFRS")

    check_cells = ["E26", "F26", "G26"]
    avo_filled = _is_any_filled(sheet_avo, check_cells)
    ifrs_filled = _is_any_filled(sheet_ifrs, check_cells)

    if avo_filled and ifrs_filled:
        msg = "Error: Both result sheets are filled out (AVO-FINMA and IFRS)"
        return False, msg, msg, None

    if not avo_filled and not ifrs_filled:
        msg = "Error: No results are filled out (AVO-FINMA or IFRS)"
        return False, msg, msg, None

    return True, "OK", "OK", (sheet_avo if avo_filled else sheet_ifrs)


def _range_has_no_empty_cells(sheet, start_col: str, end_col: str, start_row: int, end_row: int) -> bool:
    for row in range(start_row, end_row + 1):
        for col_ord in range(ord(start_col), ord(end_col) + 1):
            addr = f"{chr(col_ord)}{row}"
            v = sheet[addr].value
            if v is None or str(v).strip() == "":
                return False
    return True

def check_business_planning_filled_three_years(wb: Workbook) -> Tuple[bool, str, str]:
    ok, outcome_str, details_str, sheet = _get_filled_results_sheet(wb)
    if not ok:
        return False, outcome_str, details_str

    # Check if Zweigniederlassungs version
    is_zweigniederlassung = _is_zweigniederlassungs_version(wb)
    is_avo = _is_avo_finma_sheet(sheet.title)

    if is_zweigniederlassung:
        # Zweigniederlassungs version: E10-G20 and E23-E30
        ok_1 = _range_has_no_empty_cells(sheet, "E", "G", 10, 20)
        ok_2 = _range_has_no_empty_cells(sheet, "E", "E", 23, 30)
    elif is_avo:
        ok_1 = _range_has_no_empty_cells(sheet, "E", "G", 11, 21)
        ok_2 = _range_has_no_empty_cells(sheet, "E", "G", 24, 35)
    else:
        ok_1 = _range_has_no_empty_cells(sheet, "E", "G", 11, 23)
        ok_2 = _range_has_no_empty_cells(sheet, "E", "G", 26, 38)

    if ok_1 and ok_2:
        return True, "OK", "Geschäftsplanung ist vollständig für drei Jahre ausgefüllt"

    return False, "Prüfen", "Geschäftsplanung ist nicht vollständig für drei Jahre ausgefüllt."


def check_sst_filled_three_years(wb: Workbook) -> Tuple[bool, str, str]:
    # Check if this is a Zweigniederlassungs version
    if _is_zweigniederlassungs_version(wb):
        return False, "Kein Rating", "Kein Rating da es sich um eine Zweigniederlassung handelt"
    
    ok, outcome_str, details_str, sheet = _get_filled_results_sheet(wb)
    if not ok:
        return False, outcome_str, details_str

    is_avo = _is_avo_finma_sheet(sheet.title)

    if is_avo:
        ok = _range_has_no_empty_cells(sheet, "E", "G", 42, 45)
    else:
        ok = _range_has_no_empty_cells(sheet, "E", "G", 44, 47)

    if ok:
        return True, "OK", "SST-Daten sind vollständig für drei Jahre ausgefüllt"

    return False, "Prüfen", "SST-Daten sind nicht vollständig für drei Jahre ausgefüllt."

def check_tied_assets_filled_three_years(wb: Workbook) -> Tuple[bool, str, str]:
    ok, outcome_str, details_str, sheet = _get_filled_results_sheet(wb)
    if not ok:
        return False, outcome_str, details_str

    # Check if Zweigniederlassungs version
    is_zweigniederlassung = _is_zweigniederlassungs_version(wb)
    is_avo = _is_avo_finma_sheet(sheet.title)

    if is_zweigniederlassung:
        # Zweigniederlassungs version: E38 to G40
        ok = _range_has_no_empty_cells(sheet, "E", "G", 38, 40)
    elif is_avo:
        ok = _range_has_no_empty_cells(sheet, "E", "G", 49, 51)
    else:
        ok = _range_has_no_empty_cells(sheet, "E", "G", 51, 54)

    if ok:
        return True, "OK", "Gebundenes Vermögen ist vollständig für drei Jahre ausgefüllt"

    return False, "Prüfen", "Gebundenes Vermögen ist nicht vollständig für drei Jahre ausgefüllt."
def check_provisions_filled_three_years(wb: Workbook) -> Tuple[bool, str, str]:
    ok, outcome_str, details_str, sheet = _get_filled_results_sheet(wb)
    if not ok:
        return False, outcome_str, details_str

    # Check if Zweigniederlassungs version
    is_zweigniederlassung = _is_zweigniederlassungs_version(wb)
    is_avo = _is_avo_finma_sheet(sheet.title)

    if is_zweigniederlassung:
        # Zweigniederlassungs version: E60 to G60
        ok_range = _range_has_no_empty_cells(sheet, "E", "G", 60, 60)
    elif is_avo:
        ok_range = _range_has_no_empty_cells(sheet, "E", "G", 71, 71)
    else:
        ok_range = _range_has_no_empty_cells(sheet, "E", "G", 73, 73)

    if ok_range:
        return True, "OK", "Rückstellungen sind vollständig für drei Jahre ausgefüllt"

    return False, "Prüfen", "Rückstellungen sind nicht vollständig für drei Jahre ausgefüllt."

def check_other_perspective_filled_three_years(wb: Workbook) -> Tuple[bool, str, str]:
    ok, outcome_str, details_str, sheet = _get_filled_results_sheet(wb)
    if not ok:
        return False, outcome_str, details_str

    # Check if Zweigniederlassungs version
    is_zweigniederlassung = _is_zweigniederlassungs_version(wb)
    is_avo = _is_avo_finma_sheet(sheet.title)
    
    if is_zweigniederlassung:
        # Zweigniederlassungs version: E63 to G63, E66 to G66, E69 to G69
        row_ranges = [(63, 63), (66, 66), (69, 69)]
    else:
        shift = 0 if is_avo else 2
        row_ranges = [
            (74, 74),
            (77, 77),
            (82, 84),
            (86, 86),
            (89, 89),
            (93, 93),
            (97, 97),
        ]

    for r1, r2 in row_ranges:
        if is_zweigniederlassung:
            for row in range(r1, r2 + 1):
                e_val = sheet[f"E{row}"].value
                f_val = sheet[f"F{row}"].value
                g_val = sheet[f"G{row}"].value

                e_filled = e_val is not None and str(e_val).strip() != ""
                f_filled = f_val is not None and str(f_val).strip() != ""
                g_filled = g_val is not None and str(g_val).strip() != ""

                if e_filled and not (f_filled and g_filled):
                    return False, "Prüfen", f"Andere Perspektive (Zeile {row}): nur teilweise ausgefüllt (Spalte E ausgefüllt, aber F und/oder G fehlen)."
        else:
            for row in range(r1 + shift, r2 + shift + 1):
                e_val = sheet[f"E{row}"].value
                f_val = sheet[f"F{row}"].value
                g_val = sheet[f"G{row}"].value

                e_filled = e_val is not None and str(e_val).strip() != ""
                f_filled = f_val is not None and str(f_val).strip() != ""
                g_filled = g_val is not None and str(g_val).strip() != ""

                if e_filled and not (f_filled and g_filled):
                    return False, "Prüfen", f"Andere Perspektive (Zeile {row}): nur teilweise ausgefüllt (Spalte E ausgefüllt, aber F und/oder G fehlen)."

    return True, "OK", "Andere Perspektive ist konsistent ausgefüllt (alle Zeilen sind entweder komplett für drei Jahre ausgefüllt oder komplett leer)"



def _range_has_no_empty_cells_cols(sheet, start_col: str, end_col: str, start_row: int, end_row: int) -> bool:
    c1 = column_index_from_string(start_col)
    c2 = column_index_from_string(end_col)
    for row in range(start_row, end_row + 1):
        for c in range(c1, c2 + 1):
            addr = f"{get_column_letter(c)}{row}"
            v = sheet[addr].value
            if v is None or str(v).strip() == "":
                return False
    return True


def _scenario_cols(scenario_index_zero_based: int) -> Tuple[str, str]:
    start_idx = column_index_from_string("K") + scenario_index_zero_based * 6
    return get_column_letter(start_idx), get_column_letter(start_idx + 2)


def check_scenarios_business_planning_filled_three_years(wb: Workbook) -> Tuple[bool, str, str]:
    ok, outcome_str, details_str, results_sheet = _get_filled_results_sheet(wb)
    if not ok:
        return False, outcome_str, details_str

    mapper = SheetNameMapper(wb)
    szenarien_sheet = mapper.get_sheet("Szenarien")

    # Check if Zweigniederlassungs version
    is_zweigniederlassung = _is_zweigniederlassungs_version(wb)
    is_avo = _is_avo_finma_sheet(results_sheet.title)

    for i, type_addr in enumerate(_scenario_type_cells()):
        if (szenarien_sheet[type_addr].value or "") == "":
            continue

        start_col, end_col = _scenario_cols(i)

        if is_zweigniederlassung:
            # Zweigniederlassungs version: K10-M20 and K23 to M30
            ok_1 = _range_has_no_empty_cells_cols(results_sheet, start_col, end_col, 10, 20)
            ok_2 = _range_has_no_empty_cells_cols(results_sheet, start_col, end_col, 23, 30)
        elif is_avo:
            ok_1 = _range_has_no_empty_cells_cols(results_sheet, start_col, end_col, 11, 21)
            ok_2 = _range_has_no_empty_cells_cols(results_sheet, start_col, end_col, 24, 35)
        else:
            ok_1 = _range_has_no_empty_cells_cols(results_sheet, start_col, end_col, 11, 23)
            ok_2 = _range_has_no_empty_cells_cols(results_sheet, start_col, end_col, 26, 38)

        if not (ok_1 and ok_2):
            return False, "Prüfen", f"Geschäftsplanung für Szenarien ist nicht vollständig ausgefüllt. Szenario {i+1} (Spalten {start_col}-{end_col}) hat fehlende Werte."

    return True, "OK", "Geschäftsplanung für alle Szenarien ist vollständig für drei Jahre ausgefüllt"


def check_scenarios_sst_filled_three_years(wb: Workbook) -> Tuple[bool, str, str]:
    # Check if this is a Zweigniederlassungs version
    if _is_zweigniederlassungs_version(wb):
        return False, "Kein Rating", "Kein Rating da es sich um eine Zweigniederlassung handelt"
    
    ok, outcome_str, details_str, results_sheet = _get_filled_results_sheet(wb)
    if not ok:
        return False, outcome_str, details_str

    mapper = SheetNameMapper(wb)
    szenarien_sheet = mapper.get_sheet("Szenarien")

    is_avo = _is_avo_finma_sheet(results_sheet.title)

    for i, type_addr in enumerate(_scenario_type_cells()):
        if (szenarien_sheet[type_addr].value or "") == "":
            continue

        start_col, end_col = _scenario_cols(i)

        if is_avo:
            ok_range = _range_has_no_empty_cells_cols(results_sheet, start_col, end_col, 42, 45)
        else:
            ok_range = _range_has_no_empty_cells_cols(results_sheet, start_col, end_col, 44, 47)

        if not ok_range:
            return False, "Prüfen", f"SST-Daten für Szenarien sind nicht vollständig ausgefüllt. Szenario {i+1} (Spalten {start_col}-{end_col}) hat fehlende Werte."

    return True, "OK", "SST-Daten für alle Szenarien sind vollständig für drei Jahre ausgefüllt"


def check_scenarios_tied_assets_filled_three_years(wb: Workbook) -> Tuple[bool, str, str]:
    ok, outcome_str, details_str, results_sheet = _get_filled_results_sheet(wb)
    if not ok:
        return False, outcome_str, details_str

    mapper = SheetNameMapper(wb)
    szenarien_sheet = mapper.get_sheet("Szenarien")

    # Check if Zweigniederlassungs version
    is_zweigniederlassung = _is_zweigniederlassungs_version(wb)
    is_avo = _is_avo_finma_sheet(results_sheet.title)

    for i, type_addr in enumerate(_scenario_type_cells()):
        if (szenarien_sheet[type_addr].value or "") == "":
            continue

        start_col, end_col = _scenario_cols(i)

        if is_zweigniederlassung:
            # Zweigniederlassungs version: K38 to M40
            ok_range = _range_has_no_empty_cells_cols(results_sheet, start_col, end_col, 38, 40)
        elif is_avo:
            ok_range = _range_has_no_empty_cells_cols(results_sheet, start_col, end_col, 49, 51)
        else:
            ok_range = _range_has_no_empty_cells_cols(results_sheet, start_col, end_col, 51, 54)

        if not ok_range:
            return False, "Prüfen", f"Gebundenes Vermögen für Szenarien ist nicht vollständig ausgefüllt. Szenario {i+1} (Spalten {start_col}-{end_col}) hat fehlende Werte."

    return True, "OK", "Gebundenes Vermögen für alle Szenarien ist vollständig für drei Jahre ausgefüllt"


def check_scenarios_provisions_filled_three_years(wb: Workbook) -> Tuple[bool, str, str]:
    ok, outcome_str, details_str, results_sheet = _get_filled_results_sheet(wb)
    if not ok:
        return False, outcome_str, details_str

    mapper = SheetNameMapper(wb)
    szenarien_sheet = mapper.get_sheet("Szenarien")

    # Check if Zweigniederlassungs version
    is_zweigniederlassung = _is_zweigniederlassungs_version(wb)
    is_avo = _is_avo_finma_sheet(results_sheet.title)
    
    if is_zweigniederlassung:
        row = 60
    else:
        row = 71 if is_avo else 73

    for i, type_addr in enumerate(_scenario_type_cells()):
        if (szenarien_sheet[type_addr].value or "") == "":
            continue

        start_col, end_col = _scenario_cols(i)
        ok_range = _range_has_no_empty_cells_cols(results_sheet, start_col, end_col, row, row)

        if not ok_range:
            return False, "Prüfen", f"Rückstellungen für Szenarien sind nicht vollständig ausgefüllt. Szenario {i+1} (Spalten {start_col}-{end_col}) hat fehlende Werte."

    return True, "OK", "Rückstellungen für alle Szenarien sind vollständig für drei Jahre ausgefüllt"


def check_scenarios_other_perspective_filled_three_years(wb: Workbook) -> Tuple[bool, str, str]:
    ok, outcome_str, details_str, results_sheet = _get_filled_results_sheet(wb)
    if not ok:
        return False, outcome_str, details_str

    mapper = SheetNameMapper(wb)
    szenarien_sheet = mapper.get_sheet("Szenarien")

    # Check if Zweigniederlassungs version
    is_zweigniederlassung = _is_zweigniederlassungs_version(wb)
    is_avo = _is_avo_finma_sheet(results_sheet.title)

    if is_zweigniederlassung:
        # Zweigniederlassungs version: K63 to M63, K66 to M66, K69 to M69
        row_ranges = [(63, 63), (66, 66), (69, 69)]
        shift = 0
    else:
        shift = 0 if is_avo else 2
        row_ranges = [
            (74, 74),
            (77, 77),
            (82, 84),
            (86, 86),
            (89, 89),
            (93, 93),
            (97, 97),
        ]

    for i, type_addr in enumerate(_scenario_type_cells()):
        if (szenarien_sheet[type_addr].value or "") == "":
            continue

        c_e, c_g = _scenario_cols(i)
        c_f = get_column_letter(column_index_from_string(c_e) + 1)

        for r1, r2 in row_ranges:
            for row in range(r1 + shift, r2 + shift + 1):
                e_val = results_sheet[f"{c_e}{row}"].value
                f_val = results_sheet[f"{c_f}{row}"].value
                g_val = results_sheet[f"{c_g}{row}"].value

                e_filled = e_val is not None and str(e_val).strip() != ""
                f_filled = f_val is not None and str(f_val).strip() != ""
                g_filled = g_val is not None and str(g_val).strip() != ""

                if e_filled and not (f_filled and g_filled):
                    return False, "Prüfen", f"Andere Perspektive für Szenarien ist nur teilweise ausgefüllt. Szenario {i+1}, Zeile {row}: Spalte {c_e} ist ausgefüllt, aber {c_f} und/oder {c_g} fehlen."

    return True, "OK", "Andere Perspektive für alle Szenarien ist konsistent ausgefüllt (alle Zeilen sind entweder komplett für drei Jahre ausgefüllt oder komplett leer)"

###### Qual. & langfr. Risiken

def check_count_longterm_risks(wb: Workbook) -> Tuple[bool, str, str]:
    # Check if this is a Zweigniederlassungs version
    if _is_zweigniederlassungs_version(wb):
        return False, "Kein Rating", "Kein Rating da es sich um eine Zweigniederlassung handelt"
    
    mapper = SheetNameMapper(wb)
    sheet = mapper.get_sheet("Qual. & langfr. Risiken")
    
    if sheet is None:
        return True, "0", "Anzahl identifizierter qualitativer und langfristiger Risiken: 0 (Tabellenblatt 'Qual. & langfr. Risiken' nicht gefunden)"

    count = 0
    for row in range(25, 40):
        if (sheet[f"C{row}"].value or "") != "":
            count += 1

    count_str = str(count)
    return True, count_str, f"Anzahl identifizierter qualitativer und langfristiger Risiken: {count_str}"

def check_treatment_of_qual_risks(wb: Workbook) -> Tuple[bool, str, str]:
    # Check if this is a Zweigniederlassungs version
    if _is_zweigniederlassungs_version(wb):
        return False, "Kein Rating", "Kein Rating da es sich um eine Zweigniederlassung handelt"
    
    mapper = SheetNameMapper(wb)
    sheet = mapper.get_sheet("Qual. & langfr. Risiken")
    
    if sheet is None:
        return False, "", "Das Tabellenblatt 'Qual. & langfr. Risiken' wurde in der Arbeitsmappe nicht gefunden"

    value = str(sheet["E4"].value or "")
    
    # Extract the value and return it (e.g., "(1)", "(2)", etc.)
    # Return True if we found something, False if empty
    if value and value.strip():
        return True, value, f"Behandlung qualitativer Risiken: {value}"
    else:
        return False, "", "Behandlung qualitativer Risiken: Keine Angabe in E4"

#### Schlussfolgerungen, Dokument.

def check_orsa_dokumentation_sufficient(wb: Workbook) -> Tuple[bool, str, str]:
    mapper = SheetNameMapper(wb)
    sheet = mapper.get_sheet("Schlussfolgerungen, Dokument.")
    
    if sheet is None:
        return False, "Prüfen", "Das Tabellenblatt 'Schlussfolgerungen, Dokument.' wurde in der Arbeitsmappe nicht gefunden"

    values = [str(sheet[f"C{row}"].value or "") for row in range(24, 31)]

    if any(v.startswith("(3)") for v in values):
        result_str = "ungenügend"
        desc = "ORSA-Dokumentation ist ungenügend: Mindestens ein Kriterium ist mit '(3)  Im ORSA nicht spezifisch behandelt' bewertet"
    elif all(v.startswith("(2)") for v in values):
        result_str = "genügend"
        desc = "ORSA-Dokumentation ist genügend: Alle ausgefüllten Kriterien sind mit '(2) Durchführung und Ergebnisse spezifisch dokumentiert' bewertet"
    elif any(v.startswith("(2)") for v in values):
        result_str = "mangelhaft"
        desc = "ORSA-Dokumentation ist mangelhaft: Nur teilweise mit '(2) Durchführung und Ergebnisse spezifisch dokumentiert' bewertet"
    elif all(v.startswith("(1)") for v in values):
        result_str = "mangelhaft"
        desc = "ORSA-Dokumentation ist mangelhaft: Nur mit '(1) Behandelt aber nicht spezifisch dokumentiert' bewertet"
    else:
        result_str = "ungenügend"
        desc = "ORSA-Dokumentation muss geprüft werden: Keine gültigen Bewertungen (2)-(3) in den Zeilen C24-30 gefunden"

    return result_str == "genügend", result_str, desc


#### New Checks

def check_orsa_version(wb: Workbook) -> Tuple[bool, str, str]:
    """Check if this is a Zweigniederlassungs or Sitzgesellschaft version.
    
    Returns:
        - outcome_bool: True if Sitzgesellschaft, False if Zweigniederlassung
        - outcome_str: 'Zweigniederlassung' or 'Sitzgesellschaft'
        - description: Description of the version detected
    """
    is_zweigniederlassung = _is_zweigniederlassungs_version(wb)
    
    if is_zweigniederlassung:
        return False, "Zweigniederlassung", "ORSA-Version: Zweigniederlassung"
    else:
        return True, "Sitzgesellschaft", "ORSA-Version: Sitzgesellschaft"


def check_count_number_mitigating_measures_other_effect(wb: Workbook) -> Tuple[bool, str, str]:
    """Count (5) entries in columns E and G (rows 9-38) in Massnahmen sheet.
    
    Returns the count as "count_E / count_G"
    """
    mapper = SheetNameMapper(wb)
    sheet = mapper.get_sheet("Massnahmen")
    
    if sheet is None:
        return True, "0 / 0", "Anzahl Massnahmen mit anderer Wirkung: 0 / 0 (Tabellenblatt 'Massnahmen' nicht gefunden)"
    
    # Count (5) in column E (rows 9-38)
    count_e = 0
    for row in range(9, 39):
        value = str(sheet[f"E{row}"].value or "")
        if value.startswith("(5)"):
            count_e += 1
    
    # Count (5) in column G (rows 9-38)
    count_g = 0
    for row in range(9, 39):
        value = str(sheet[f"G{row}"].value or "")
        if value.startswith("(5)"):
            count_g += 1
    
    result_str = f"{count_e} / {count_g}"
    return True, result_str, f"Anzahl Massnahmen mit anderer Wirkung (Spalte E / Spalte G): {result_str}"


def check_count_number_mitigating_measures_risk_accepted(wb: Workbook) -> Tuple[bool, str, str]:
    """Count (6) entries in columns E and G (rows 9-38) in Massnahmen sheet.
    
    Returns the count as "count_E / count_G"
    """
    mapper = SheetNameMapper(wb)
    sheet = mapper.get_sheet("Massnahmen")
    
    if sheet is None:
        return True, "0 / 0", "Anzahl akzeptierter Risiken: 0 / 0 (Tabellenblatt 'Massnahmen' nicht gefunden)"
    
    # Count (6) in column E (rows 9-38)
    count_e = 0
    for row in range(9, 39):
        value = str(sheet[f"E{row}"].value or "")
        if value.startswith("(6)"):
            count_e += 1
    
    # Count (6) in column G (rows 9-38)
    count_g = 0
    for row in range(9, 39):
        value = str(sheet[f"G{row}"].value or "")
        if value.startswith("(6)"):
            count_g += 1
    
    result_str = f"{count_e} / {count_g}"
    return True, result_str, f"Anzahl akzeptierter Risiken (Spalte E / Spalte G): {result_str}"


def check_count_number_other_measures_other_effect(wb: Workbook) -> Tuple[bool, str, str]:
    """Count (4) entries in column F rows 44-53 and 57-66 in Massnahmen sheet.
    
    Returns the total count.
    """
    mapper = SheetNameMapper(wb)
    sheet = mapper.get_sheet("Massnahmen")
    
    if sheet is None:
        return True, "0", "Anzahl sonstiger Massnahmen mit anderer Wirkung: 0 (Tabellenblatt 'Massnahmen' nicht gefunden)"
    
    count = 0
    # Count (4) in column F rows 44-53
    for row in range(44, 54):
        value = str(sheet[f"F{row}"].value or "")
        if value.startswith("(4)"):
            count += 1
    
    # Count (4) in column F rows 57-66
    for row in range(57, 67):
        value = str(sheet[f"F{row}"].value or "")
        if value.startswith("(4)"):
            count += 1
    
    count_str = str(count)
    return True, count_str, f"Anzahl sonstiger Massnahmen mit anderer Wirkung: {count_str}"


def check_liquidity_filled_three_years(wb: Workbook) -> Tuple[bool, str, str]:
    """Check if liquidity is filled for three years (base case).
    
    For AVO-FINMA: line 86
    For IFRS: line 88
    For Zweigniederlassung: line 66
    """
    mapper = SheetNameMapper(wb)
    is_zweigniederlassung = _is_zweigniederlassungs_version(wb)
    
    if is_zweigniederlassung:
        # Zweigniederlassung version - check line 66 in Ergebnisse sheet
        sheet = mapper.get_sheet("Ergebnisse")
        if sheet is None:
            return False, "Nein", "Tabellenblatt 'Ergebnisse' nicht gefunden"
        
        row = 66
        e_val = sheet[f"E{row}"].value
        f_val = sheet[f"F{row}"].value
        g_val = sheet[f"G{row}"].value
        
        all_filled = all(val is not None and str(val).strip() != "" for val in [e_val, f_val, g_val])
        result_str = "Ja" if all_filled else "Nein"
        
        if all_filled:
            return True, result_str, f"Liquidität für drei Jahre ausgefüllt (Zeile {row}): Ja"
        else:
            return False, result_str, f"Liquidität für drei Jahre ausgefüllt (Zeile {row}): Nein - nicht alle drei Jahre sind ausgefüllt"
    else:
        # Sitzgesellschaft version - check both AVO-FINMA (line 86) and IFRS (line 88)
        sheet_avo = mapper.get_sheet("Ergebnisse_AVO-FINMA")
        sheet_ifrs = mapper.get_sheet("Ergebnisse_IFRS")
        
        results = []
        
        # Check AVO-FINMA line 86
        if sheet_avo is not None:
            row = 86
            e_val = sheet_avo[f"E{row}"].value
            f_val = sheet_avo[f"F{row}"].value
            g_val = sheet_avo[f"G{row}"].value
            avo_filled = all(val is not None and str(val).strip() != "" for val in [e_val, f_val, g_val])
            results.append(("AVO-FINMA", avo_filled, row))
        else:
            results.append(("AVO-FINMA", False, 86))
        
        # Check IFRS line 88
        if sheet_ifrs is not None:
            row = 88
            e_val = sheet_ifrs[f"E{row}"].value
            f_val = sheet_ifrs[f"F{row}"].value
            g_val = sheet_ifrs[f"G{row}"].value
            ifrs_filled = all(val is not None and str(val).strip() != "" for val in [e_val, f_val, g_val])
            results.append(("IFRS", ifrs_filled, row))
        else:
            results.append(("IFRS", False, 88))
        
        # Both must be filled
        all_filled = all(filled for _, filled, _ in results)
        result_str = "Ja" if all_filled else "Nein"
        
        desc_parts = [f"{name} Zeile {row}: {'Ja' if filled else 'Nein'}" for name, filled, row in results]
        desc = f"Liquidität für drei Jahre ausgefüllt - {', '.join(desc_parts)}"
        
        return all_filled, result_str, desc


def check_scenarios_liquidity_filled_three_years(wb: Workbook) -> Tuple[bool, str, str]:
    """Check if scenario liquidity is filled for three years.
    
    For AVO-FINMA: line 86 in scenarios sheet
    For IFRS: line 88 in scenarios sheet
    For Zweigniederlassung: line 66 in scenarios sheet
    """
    mapper = SheetNameMapper(wb)
    is_zweigniederlassung = _is_zweigniederlassungs_version(wb)
    
    if is_zweigniederlassung:
        # Zweigniederlassung version - check line 66 in Szenarien sheet
        sheet = mapper.get_sheet("Szenarien")
        if sheet is None:
            return False, "Nein", "Tabellenblatt 'Szenarien' nicht gefunden"
        
        row = 66
        e_val = sheet[f"E{row}"].value
        f_val = sheet[f"F{row}"].value
        g_val = sheet[f"G{row}"].value
        
        all_filled = all(val is not None and str(val).strip() != "" for val in [e_val, f_val, g_val])
        result_str = "Ja" if all_filled else "Nein"
        
        if all_filled:
            return True, result_str, f"Szenarien Liquidität für drei Jahre ausgefüllt (Zeile {row}): Ja"
        else:
            return False, result_str, f"Szenarien Liquidität für drei Jahre ausgefüllt (Zeile {row}): Nein - nicht alle drei Jahre sind ausgefüllt"
    else:
        # Sitzgesellschaft version - check both AVO-FINMA (line 86) and IFRS (line 88) in Szenarien sheets
        sheet_avo = mapper.get_sheet("Szenarien_AVO-FINMA")
        sheet_ifrs = mapper.get_sheet("Szenarien_IFRS")
        
        results = []
        
        # Check AVO-FINMA line 86
        if sheet_avo is not None:
            row = 86
            e_val = sheet_avo[f"E{row}"].value
            f_val = sheet_avo[f"F{row}"].value
            g_val = sheet_avo[f"G{row}"].value
            avo_filled = all(val is not None and str(val).strip() != "" for val in [e_val, f_val, g_val])
            results.append(("AVO-FINMA", avo_filled, row))
        else:
            results.append(("AVO-FINMA", False, 86))
        
        # Check IFRS line 88
        if sheet_ifrs is not None:
            row = 88
            e_val = sheet_ifrs[f"E{row}"].value
            f_val = sheet_ifrs[f"F{row}"].value
            g_val = sheet_ifrs[f"G{row}"].value
            ifrs_filled = all(val is not None and str(val).strip() != "" for val in [e_val, f_val, g_val])
            results.append(("IFRS", ifrs_filled, row))
        else:
            results.append(("IFRS", False, 88))
        
        # Both must be filled
        all_filled = all(filled for _, filled, _ in results)
        result_str = "Ja" if all_filled else "Nein"
        
        desc_parts = [f"{name} Zeile {row}: {'Ja' if filled else 'Nein'}" for name, filled, row in results]
        desc = f"Szenarien Liquidität für drei Jahre ausgefüllt - {', '.join(desc_parts)}"
        
        return all_filled, result_str, desc


##########################



REGISTERED_CHECKS: list[Tuple[str, CheckFunction]] = [
    ("check_responsible_person", check_responsible_person),
    ("check_data_recency_geschaeftsplanung", check_data_recency_geschaeftsplanung),
    ("check_data_recency_risikoidentifikation",check_data_recency_risikoidentifikation,),
    ("check_data_recency_szenarien", check_data_recency_szenarien),
    ("check_board_approved_orsa", check_board_approved_orsa),
    ("check_risikobeurteilung_method", check_risikobeurteilung_method),
    ("check_risk_criteria_sufficient", check_risk_criteria_sufficient),
    ("check_finanzmarktrisiko_count", check_finanzmarktrisiko_count),
    ("check_versicherungsrisiko_count", check_versicherungsrisiko_count),
    ("check_kreditrisiko_count", check_kreditrisiko_count),
    ("check_liquiditaetsrisiko_count", check_liquiditaetsrisiko_count),
    ("check_operationelles_risiko_count", check_operationelles_risiko_count),
    ("check_strategisches_umfeld_risiko_count",check_strategisches_umfeld_risiko_count,),
    ("check_anderes_risiko_count", check_anderes_risiko_count),
    ("check_count_number_mitigating_measures", check_count_number_mitigating_measures),
    ("check_count_number_potential_mitigating_measures",check_count_number_potential_mitigating_measures,),
    ("check_count_other_measures", check_count_other_measures),
    ("check_count_potential_other_measures", check_count_potential_other_measures),
    ("check_risks_are_all_mitigated", check_risks_are_all_mitigated),
    ("check_any_nonmitigating_measures", check_any_nonmitigating_measures),
    ("check_count_all_scenarios", check_count_all_scenarios),
    ("check_count_advers_scenarios", check_count_advers_scenarios),
    ("check_count_existenzbedrohend_scenarios",check_count_existenzbedrohend_scenarios,),
    ("check_count_other_scenarios", check_count_other_scenarios),
    ("check_every_scenario_has_event", check_every_scenario_has_event),
    ("check_count_scenarios_only_one_event", check_count_scenarios_only_one_event),
    ("check_count_scenarios_multiple_events", check_count_scenarios_multiple_events),
    ("check_every_scenario_has_risk", check_every_scenario_has_risk),
    ("check_count_scenarios_only_one_risk", check_count_scenarios_only_one_risk),
    ("check_count_scenrios_multiple_risks", check_count_scenrios_multiple_risks),
    ("check_business_planning_filled_three_years", check_business_planning_filled_three_years),
    ("check_sst_filled_three_years", check_sst_filled_three_years),
    ("check_tied_assets_filled_three_years", check_tied_assets_filled_three_years),
    ("check_provisions_filled_three_years", check_provisions_filled_three_years),
    ("check_other_perspective_filled_three_years", check_other_perspective_filled_three_years),
    ("check_scenarios_business_planning_filled_three_years", check_scenarios_business_planning_filled_three_years),
    ("check_scenarios_sst_filled_three_years", check_scenarios_sst_filled_three_years),
    ("check_scenarios_tied_assets_filled_three_years", check_scenarios_tied_assets_filled_three_years),
    ("check_scenarios_provisions_filled_three_years", check_scenarios_provisions_filled_three_years),
    ("check_scenarios_other_perspective_filled_three_years", check_scenarios_other_perspective_filled_three_years),
    ("check_count_longterm_risks", check_count_longterm_risks),
    ("check_treatment_of_qual_risks", check_treatment_of_qual_risks),
    ("check_orsa_dokumentation_sufficient", check_orsa_dokumentation_sufficient),
    ("check_orsa_version", check_orsa_version),
    ("check_count_number_mitigating_measures_other_effect", check_count_number_mitigating_measures_other_effect),
    ("check_count_number_mitigating_measures_risk_accepted", check_count_number_mitigating_measures_risk_accepted),
    ("check_count_number_other_measures_other_effect", check_count_number_other_measures_other_effect),
    ("check_liquidity_filled_three_years", check_liquidity_filled_three_years),
    ("check_scenarios_liquidity_filled_three_years", check_scenarios_liquidity_filled_three_years),
]


def get_all_checks() -> list[Tuple[str, CheckFunction]]:
    """Get all registered check functions.

    Returns:
        List of tuples (check_name, check_function)
    """
    return REGISTERED_CHECKS.copy()


def run_check(
    check_name: str, check_function: CheckFunction, workbook: Workbook
) -> Tuple[bool, str, str]:
    """Execute a single check with error handling.

    Args:
        check_name: Name of the check
        check_function: Check function to execute
        workbook: Workbook to check

    Returns:
        Tuple of (outcome, outcome_str, description)
    """
    try:
        return check_function(workbook)
    except Exception as e:
        logger.error(f"Check '{check_name}' failed with error: {e}")
        return False, "zu prüfen", f"Check failed with error: {str(e)}"
