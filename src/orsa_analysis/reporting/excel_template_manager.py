"""Template manager for Excel report generation."""

import logging
from pathlib import Path
from typing import Any, Union

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

logger = logging.getLogger(__name__)


class ExcelTemplateManager:
    """Manages Excel template operations for report generation.
    
    This manager loads the template file directly and modifies it in place.
    This approach preserves all formatting including conditional formatting,
    data validation, and other advanced Excel features that would be lost
    when copying sheets cell by cell.
    """

    def __init__(self, template_path: Path):
        """Initialize template manager.

        Args:
            template_path: Path to the Excel template file

        Raises:
            FileNotFoundError: If template file doesn't exist
        """
        if not template_path.exists():
            raise FileNotFoundError(f"Template file not found: {template_path}")

        self.template_path = template_path
        self.output_wb = None
        logger.info(f"Template manager initialized with: {template_path}")

    def create_output_workbook(self, source_path: Path) -> Workbook:
        """Load template workbook for modification.

        This method loads the template file directly. The workbook can then be
        modified and saved to a different location, preserving all formatting.
        The source file path is validated but its content is not included in the output.

        Args:
            source_path: Path to source ORSA Excel file (for validation only)

        Returns:
            Template workbook ready for modification
            
        Raises:
            FileNotFoundError: If source file doesn't exist
        """
        if not source_path.exists():
            raise FileNotFoundError(f"Source file not found: {source_path}")

        # Load template file directly - this preserves all formatting including
        # conditional formatting, data validation, and other advanced features
        self.output_wb = load_workbook(self.template_path)
        logger.debug(f"Loaded template with sheets: {self.output_wb.sheetnames}")

        logger.info(
            f"Loaded template workbook with {len(self.output_wb.sheetnames)} sheet(s)"
        )
        return self.output_wb

    @staticmethod
    def _convert_numeric_string(value: Any) -> Union[int, float, Any]:
        """Convert string representations of numbers to actual numeric types.
        
        This ensures that numeric values stored as strings are converted to
        actual numbers before being written to Excel cells. This is important
        because Excel formulas work correctly with numeric types but may fail
        with string representations of numbers.
        
        Args:
            value: The value to potentially convert
            
        Returns:
            - int: If value is a string representing an integer (e.g., "42")
            - float: If value is a string representing a decimal (e.g., "42.5")
            - original value: If value is not a numeric string or already a number
            
        Examples:
            >>> _convert_numeric_string("42")
            42
            >>> _convert_numeric_string("42.5")
            42.5
            >>> _convert_numeric_string("genügend")
            "genügend"
            >>> _convert_numeric_string(42)
            42
        """
        # Only process string values
        if not isinstance(value, str):
            return value
        
        # Skip empty strings
        if not value or value.isspace():
            return value
        
        # Try to convert to numeric type
        try:
            # Check if it's an integer (no decimal point)
            if '.' not in value:
                # Try converting to int
                return int(value)
            else:
                # Try converting to float
                return float(value)
        except (ValueError, TypeError):
            # If conversion fails, return original value
            return value

    def write_cell_value(
        self, sheet_name: str, cell_address: str, value: Any
    ) -> bool:
        """Write a value to a specific cell in the output workbook.
        
        String representations of numbers are automatically converted to numeric
        types before writing to ensure Excel formulas work correctly.

        Args:
            sheet_name: Name of the worksheet
            cell_address: Cell address (e.g., "A1", "B5")
            value: Value to write (numeric strings are auto-converted)

        Returns:
            True if successful, False otherwise
        """
        if self.output_wb is None:
            logger.error("Output workbook not created yet")
            return False

        if sheet_name not in self.output_wb.sheetnames:
            logger.error(f"Sheet not found: {sheet_name}")
            return False

        try:
            # Convert numeric strings to actual numbers
            converted_value = self._convert_numeric_string(value)
            
            sheet = self.output_wb[sheet_name]
            sheet[cell_address] = converted_value
            
            # Log conversion if it occurred
            if converted_value != value and isinstance(value, str):
                logger.debug(
                    f"Converted string '{value}' to {type(converted_value).__name__} "
                    f"{converted_value} for {sheet_name}!{cell_address}"
                )
            else:
                logger.debug(f"Wrote value to {sheet_name}!{cell_address}: {converted_value}")
            
            return True
        except Exception as e:
            logger.error(f"Error writing to {sheet_name}!{cell_address}: {e}")
            return False

    def save_workbook(self, output_path: Path) -> None:
        """Save the output workbook to file.

        Args:
            output_path: Path where to save the workbook
        """
        if self.output_wb is None:
            logger.error("No output workbook to save")
            return

        # Create output directory if needed
        output_path.parent.mkdir(parents=True, exist_ok=True)

        self.output_wb.save(output_path)
        logger.info(f"Saved output workbook to: {output_path}")

    def close(self) -> None:
        """Close all open workbooks."""
        self.output_wb = None
        logger.debug("Closed template manager workbooks")
