"""Template and Excel file management for report generation."""

import logging
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


logger = logging.getLogger(__name__)


class TemplateManager:
    """Manage template and source file operations for report generation."""
    
    def __init__(self, template_path: Path):
        """Initialize template manager.
        
        Args:
            template_path: Path to template Excel file
            
        Raises:
            FileNotFoundError: If template file does not exist
        """
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template file not found: {template_path}")
        
        logger.debug(f"TemplateManager initialized with template: {template_path}")
    
    def load_template(self) -> Workbook:
        """Load template workbook.
        
        Returns:
            Loaded workbook
            
        Raises:
            Exception: If template cannot be loaded
        """
        try:
            wb = openpyxl.load_workbook(self.template_path)
            logger.debug(f"Loaded template with {len(wb.sheetnames)} sheets: {wb.sheetnames}")
            return wb
        except Exception as e:
            logger.error(f"Failed to load template: {e}")
            raise
    
    def load_source_file(self, source_path: Path) -> Optional[Workbook]:
        """Load source ORSA workbook.
        
        Args:
            source_path: Path to source Excel file
            
        Returns:
            Loaded workbook, or None if file cannot be loaded
        """
        if not source_path.exists():
            logger.warning(f"Source file not found: {source_path}")
            return None
        
        try:
            wb = openpyxl.load_workbook(source_path, data_only=True)
            logger.debug(f"Loaded source file with {len(wb.sheetnames)} sheets")
            return wb
        except Exception as e:
            logger.error(f"Failed to load source file {source_path}: {e}")
            return None
    
    def create_output_workbook(self, 
                               template_wb: Workbook,
                               source_wb: Optional[Workbook] = None) -> Workbook:
        """Create output workbook from template and optional source.
        
        The output workbook will have:
        1. First: All sheets from template (especially 'Auswertung')
        2. Then: All sheets from source file (if provided)
        
        Args:
            template_wb: Loaded template workbook
            source_wb: Optional source ORSA workbook
            
        Returns:
            New workbook for output
        """
        # Create new workbook
        output_wb = Workbook()
        
        # Remove default sheet
        if "Sheet" in output_wb.sheetnames:
            output_wb.remove(output_wb["Sheet"])
        
        # Copy all sheets from template first
        logger.debug("Copying template sheets to output workbook")
        for sheet_name in template_wb.sheetnames:
            self._copy_sheet(template_wb[sheet_name], output_wb, sheet_name)
        
        # Then copy all sheets from source if provided
        if source_wb is not None:
            logger.debug("Copying source sheets to output workbook")
            for sheet_name in source_wb.sheetnames:
                # Avoid name conflicts by appending suffix if needed
                target_name = sheet_name
                if sheet_name in output_wb.sheetnames:
                    target_name = f"{sheet_name}_Source"
                    logger.debug(f"Sheet name conflict: renaming '{sheet_name}' to '{target_name}'")
                
                self._copy_sheet(source_wb[sheet_name], output_wb, target_name)
        
        logger.info(f"Created output workbook with {len(output_wb.sheetnames)} sheets")
        return output_wb
    
    def _copy_sheet(self, source_sheet: Worksheet, 
                   target_wb: Workbook, 
                   target_name: str) -> Worksheet:
        """Copy a worksheet to target workbook.
        
        Args:
            source_sheet: Sheet to copy from
            target_wb: Workbook to copy to
            target_name: Name for the new sheet
            
        Returns:
            Created worksheet
        """
        target_sheet = target_wb.create_sheet(title=target_name)
        
        # Copy cell values and basic formatting
        for row in source_sheet.iter_rows():
            for cell in row:
                target_cell = target_sheet[cell.coordinate]
                target_cell.value = cell.value
                
                # Copy basic formatting if it exists
                if cell.has_style:
                    target_cell.font = cell.font.copy()
                    target_cell.border = cell.border.copy()
                    target_cell.fill = cell.fill.copy()
                    target_cell.number_format = cell.number_format
                    target_cell.alignment = cell.alignment.copy()
        
        # Copy column dimensions
        for col_letter, dimension in source_sheet.column_dimensions.items():
            target_sheet.column_dimensions[col_letter].width = dimension.width
        
        # Copy row dimensions
        for row_num, dimension in source_sheet.row_dimensions.items():
            target_sheet.row_dimensions[row_num].height = dimension.height
        
        # Copy merged cells
        for merged_cell_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_cell_range))
        
        logger.debug(f"Copied sheet '{source_sheet.title}' as '{target_name}'")
        return target_sheet
    
    def write_cell_value(self, wb: Workbook, 
                        sheet_name: str,
                        cell_address: str, 
                        value: any) -> bool:
        """Write value to specific cell.
        
        Args:
            wb: Workbook to modify
            sheet_name: Name of sheet containing the cell
            cell_address: Excel cell address (e.g., "C8")
            value: Value to write
            
        Returns:
            True if successful, False otherwise
        """
        try:
            if sheet_name not in wb.sheetnames:
                logger.error(f"Sheet '{sheet_name}' not found in workbook")
                return False
            
            sheet = wb[sheet_name]
            sheet[cell_address] = value
            logger.debug(f"Wrote value to {sheet_name}!{cell_address}: {value}")
            return True
        
        except Exception as e:
            logger.error(f"Failed to write to {sheet_name}!{cell_address}: {e}")
            return False
    
    def save_workbook(self, wb: Workbook, output_path: Path) -> bool:
        """Save workbook to file.
        
        Args:
            wb: Workbook to save
            output_path: Path where to save the file
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Ensure output directory exists
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"Saved workbook to: {output_path}")
            return True
        
        except Exception as e:
            logger.error(f"Failed to save workbook to {output_path}: {e}")
            return False
    
    def close_workbook(self, wb: Workbook) -> None:
        """Close workbook to free resources.
        
        Args:
            wb: Workbook to close
        """
        try:
            wb.close()
            logger.debug("Closed workbook")
        except Exception as e:
            logger.warning(f"Error closing workbook: {e}")
