"""Excel file reader module using openpyxl."""

from pathlib import Path
from typing import Optional
import logging

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

logger = logging.getLogger(__name__)


class ExcelReader:
    """Handles reading Excel files using openpyxl."""

    def __init__(self, data_only: bool = True, read_only: bool = False):
        """Initialize the Excel reader.

        Args:
            data_only: If True, cell values are read instead of formulas
            read_only: If True, file is opened in read-only mode for better performance
        """
        self.data_only = data_only
        self.read_only = read_only

    def load_file(self, file_path: Path) -> Optional[Workbook]:
        """Load an Excel file and return a Workbook object.

        Args:
            file_path: Path to the Excel file

        Returns:
            Workbook object if successful, None if loading failed

        Raises:
            FileNotFoundError: If the file does not exist
            ValueError: If the file is not a valid Excel file
        """
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        if not file_path.suffix.lower() in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
            raise ValueError(f"Invalid Excel file extension: {file_path.suffix}")

        try:
            workbook = load_workbook(
                filename=str(file_path),
                data_only=self.data_only,
                read_only=self.read_only,
            )
            logger.info(f"Successfully loaded workbook: {file_path.name}")
            return workbook
        except Exception as e:
            logger.error(f"Failed to load {file_path}: {e}")
            raise

    def get_sheet_names(self, workbook: Workbook) -> list[str]:
        """Get all sheet names from a workbook.

        Args:
            workbook: Workbook object

        Returns:
            List of sheet names
        """
        return workbook.sheetnames

    def close_workbook(self, workbook: Workbook) -> None:
        """Close a workbook to free resources.

        Args:
            workbook: Workbook object to close
        """
        if workbook:
            workbook.close()
