"""Tests for multi-language sheet name mapping."""

import pytest
from openpyxl import Workbook

from orsa_analysis.checks.sheet_mapper import SheetNameMapper, SHEET_NAME_MAPPING


class TestSheetNameMapper:
    """Test the SheetNameMapper class."""

    def test_detect_german_workbook(self):
        """Test language detection for German workbook."""
        wb = Workbook()
        wb.remove(wb.active)
        
        # Add German sheet names
        wb.create_sheet("Auswertung")
        wb.create_sheet("Risiken")
        wb.create_sheet("Ergebnisse_AVO-FINMA")
        
        mapper = SheetNameMapper(wb)
        assert mapper.detected_language == "DE"

    def test_detect_english_workbook(self):
        """Test language detection for English workbook."""
        wb = Workbook()
        wb.remove(wb.active)
        
        # Add English sheet names
        wb.create_sheet("General details")
        wb.create_sheet("Measures")
        wb.create_sheet("Results_ISO-FINMA")
        
        mapper = SheetNameMapper(wb)
        assert mapper.detected_language == "EN"

    def test_detect_french_workbook(self):
        """Test language detection for French workbook."""
        wb = Workbook()
        wb.remove(wb.active)
        
        # Add French sheet names
        wb.create_sheet("Info. générales")
        wb.create_sheet("Mesures")
        wb.create_sheet("Résultats_OS-FINMA")
        
        mapper = SheetNameMapper(wb)
        assert mapper.detected_language == "FR"

    def test_get_sheet_name_german(self):
        """Test getting sheet name from German reference in German workbook."""
        wb = Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Ergebnisse_AVO-FINMA")
        
        mapper = SheetNameMapper(wb)
        assert mapper.get_sheet_name("Ergebnisse_AVO-FINMA") == "Ergebnisse_AVO-FINMA"

    def test_get_sheet_name_english(self):
        """Test getting sheet name from German reference in English workbook."""
        wb = Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Results_ISO-FINMA")
        
        mapper = SheetNameMapper(wb)
        assert mapper.get_sheet_name("Ergebnisse_AVO-FINMA") == "Results_ISO-FINMA"

    def test_get_sheet_name_french(self):
        """Test getting sheet name from German reference in French workbook."""
        wb = Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Résultats_OS-FINMA")
        
        mapper = SheetNameMapper(wb)
        assert mapper.get_sheet_name("Ergebnisse_AVO-FINMA") == "Résultats_OS-FINMA"

    def test_get_sheet_name_not_found(self):
        """Test getting sheet name that doesn't exist."""
        wb = Workbook()
        wb.remove(wb.active)
        wb.create_sheet("SomeSheet")
        
        mapper = SheetNameMapper(wb)
        assert mapper.get_sheet_name("NonExistentSheet") is None

    def test_get_sheet_object_german(self):
        """Test getting sheet object in German workbook."""
        wb = Workbook()
        wb.remove(wb.active)
        sheet = wb.create_sheet("Ergebnisse_AVO-FINMA")
        sheet["A1"] = "Test"
        
        mapper = SheetNameMapper(wb)
        result_sheet = mapper.get_sheet("Ergebnisse_AVO-FINMA")
        
        assert result_sheet is not None
        assert result_sheet["A1"].value == "Test"

    def test_get_sheet_object_english(self):
        """Test getting sheet object in English workbook."""
        wb = Workbook()
        wb.remove(wb.active)
        sheet = wb.create_sheet("Results_ISO-FINMA")
        sheet["B2"] = "English"
        
        mapper = SheetNameMapper(wb)
        result_sheet = mapper.get_sheet("Ergebnisse_AVO-FINMA")
        
        assert result_sheet is not None
        assert result_sheet["B2"].value == "English"

    def test_has_sheet_true(self):
        """Test has_sheet returns True for existing sheet."""
        wb = Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Ergebnisse_AVO-FINMA")
        
        mapper = SheetNameMapper(wb)
        assert mapper.has_sheet("Ergebnisse_AVO-FINMA") is True

    def test_has_sheet_false(self):
        """Test has_sheet returns False for non-existing sheet."""
        wb = Workbook()
        wb.remove(wb.active)
        wb.create_sheet("SomeSheet")
        
        mapper = SheetNameMapper(wb)
        assert mapper.has_sheet("Ergebnisse_AVO-FINMA") is False

    def test_get_all_mapped_sheets_german(self):
        """Test getting all mapped sheets in German workbook."""
        wb = Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Auswertung")
        wb.create_sheet("Risiken")
        wb.create_sheet("Ergebnisse_AVO-FINMA")
        
        mapper = SheetNameMapper(wb)
        mapped = mapper.get_all_mapped_sheets()
        
        assert "Auswertung" in mapped
        assert "Risiken" in mapped
        assert "Ergebnisse_AVO-FINMA" in mapped
        assert mapped["Auswertung"] == "Auswertung"
        assert mapped["Risiken"] == "Risiken"

    def test_get_all_mapped_sheets_english(self):
        """Test getting all mapped sheets in English workbook."""
        wb = Workbook()
        wb.remove(wb.active)
        wb.create_sheet("General details")
        wb.create_sheet("Measures")
        wb.create_sheet("Results_ISO-FINMA")
        
        mapper = SheetNameMapper(wb)
        mapped = mapper.get_all_mapped_sheets()
        
        assert "Auswertung" in mapped
        assert "Risiken" in mapped
        assert "Ergebnisse_AVO-FINMA" in mapped
        assert mapped["Auswertung"] == "General details"
        assert mapped["Risiken"] == "Measures"

    def test_mapping_completeness(self):
        """Test that all expected German sheet names are in mapping."""
        expected_sheets = [
            "Mgmt. Summary",
            "Auswertung",
            "Allgem. Angaben",
            "Risiken",
            "Massnahmen",
            "Szenarien",
            "Ergebnisse_AVO-FINMA",
            "Ergebnisse_IFRS",
            "Qual. & langfr. Risiken",
            "Schlussfolgerungen, Dokument.",
            "Drop-downs",
        ]
        
        for sheet_name in expected_sheets:
            assert sheet_name in SHEET_NAME_MAPPING, f"Missing mapping for {sheet_name}"
            assert "EN" in SHEET_NAME_MAPPING[sheet_name]
            assert "FR" in SHEET_NAME_MAPPING[sheet_name]
