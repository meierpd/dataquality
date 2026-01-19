"""Meta-level tests for the rules module.

This module tests the check framework itself, not individual check implementations.
Individual checks are subject to change, so we only test:
- That we can retrieve registered checks
- That check execution works correctly
- That the check interface is consistent
"""

import pytest
from openpyxl import Workbook

from orsa_analysis.checks.rules import (
    get_all_checks,
    run_check,
)


@pytest.fixture
def basic_workbook():
    """Create a basic workbook for testing the check framework."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Header1"
    ws["B1"] = "Header2"
    ws["A2"] = "Data1"
    ws["B2"] = "Data2"
    return wb


class TestGetAllChecks:
    """Test cases for retrieving registered checks."""

    def test_get_all_checks_returns_list(self):
        """Test that get_all_checks returns a list."""
        checks = get_all_checks()
        assert isinstance(checks, list)
        assert len(checks) > 0

    def test_get_all_checks_format(self):
        """Test that each check is a tuple of (name, function)."""
        checks = get_all_checks()

        for check in checks:
            assert isinstance(check, tuple)
            assert len(check) == 2
            assert isinstance(check[0], str)
            assert callable(check[1])

    def test_get_all_checks_has_content(self):
        """Test that there are checks registered."""
        checks = get_all_checks()
        assert len(checks) >= 3, "Should have at least a few checks registered"

    def test_check_names_are_unique(self):
        """Test that all check names are unique."""
        checks = get_all_checks()
        check_names = [name for name, _ in checks]
        assert len(check_names) == len(set(check_names)), "Check names should be unique"


class TestRunCheck:
    """Test cases for the run_check wrapper function."""

    def test_run_check_success(self, basic_workbook):
        """Test successful check execution returns correct types."""
        checks = get_all_checks()
        assert len(checks) > 0, "Need at least one check to test"
        
        check_name, check_function = checks[0]
        outcome, outcome_str, description = run_check(
            check_name, check_function, basic_workbook
        )

        assert isinstance(outcome, bool)
        assert isinstance(outcome_str, str)
        assert isinstance(description, str)

    def test_run_check_with_error(self, basic_workbook):
        """Test check execution with error handling."""

        def failing_check(wb):
            raise ValueError("Test error")

        outcome, outcome_str, description = run_check(
            "failing_check", failing_check, basic_workbook
        )

        assert outcome is False
        assert outcome_str == "zu prüfen"
        assert "error" in description.lower()

    def test_run_check_all_registered(self, basic_workbook):
        """Test running all registered checks works and returns correct format."""
        checks = get_all_checks()

        for check_name, check_function in checks:
            outcome, outcome_str, description = run_check(
                check_name, check_function, basic_workbook
            )

            # Verify return types
            assert isinstance(outcome, bool), f"Check '{check_name}' should return bool as first value"
            assert isinstance(outcome_str, str), f"Check '{check_name}' should return str as second value"
            assert isinstance(description, str), f"Check '{check_name}' should return str as third value"
            
            # Verify description is not empty
            assert len(description) > 0, f"Check '{check_name}' should return non-empty description"

    def test_run_check_consistent_interface(self, basic_workbook):
        """Test that all checks follow the same interface pattern."""
        checks = get_all_checks()

        for check_name, check_function in checks:
            # Should be able to call with just a workbook
            try:
                result = check_function(basic_workbook)
                assert isinstance(result, tuple), f"Check '{check_name}' should return a tuple"
                assert len(result) == 3, f"Check '{check_name}' should return a 3-tuple"
            except Exception as e:
                pytest.fail(f"Check '{check_name}' raised unexpected exception: {e}")
