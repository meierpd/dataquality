"""Unit tests for the orchestrator module."""

import pytest
from pathlib import Path
from openpyxl import Workbook
from typing import List, Tuple, Dict, Any

from orsa_analysis.core.orchestrator import (
    ORSAPipeline,
    CachedDocumentProcessor,
)
from orsa_analysis.core.database_manager import CheckResult


class MockDatabaseManager:
    """Mock DatabaseManager for testing."""
    
    def __init__(self):
        self.stored_results = []
        self.existing_versions = []
        self._versions_by_institute = {}
    
    def write_results(self, results: List[CheckResult]) -> int:
        self.stored_results.extend(results)
        # Track versions
        for result in results:
            key = (result.institute_id, result.file_hash)
            if key not in self._versions_by_institute:
                self._versions_by_institute[key] = {
                    "institute_id": result.institute_id,
                    "file_hash": result.file_hash,
                    "version_number": result.version_number,
                }
                self.existing_versions.append(self._versions_by_institute[key])
        return len(results)
    
    def get_existing_versions(self) -> List[Dict[str, Any]]:
        return self.existing_versions.copy()
    
    def close(self):
        pass


class MockORSADocumentSourcer:
    """Mock ORSADocumentSourcer for testing."""
    
    def __init__(self, documents: List[Tuple[str, Path, str]]):
        self.documents = documents
    
    def load(self) -> List[Tuple[str, Path, str]]:
        return self.documents


@pytest.fixture
def sample_excel_file(tmp_path):
    """Create a sample Excel file for testing."""
    file_path = tmp_path / "INST001_report.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    ws["A1"] = "Header1"
    ws["B1"] = "Header2"
    ws["A2"] = "Data1"
    ws["B2"] = "Data2"
    wb.save(file_path)
    return file_path


@pytest.fixture
def sample_excel_file2(tmp_path):
    """Create another sample Excel file for testing."""
    file_path = tmp_path / "INST002_report.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Col1"
    ws["B1"] = "Col2"
    ws["A2"] = "Value1"
    ws["B2"] = "Value2"
    wb.save(file_path)
    return file_path


@pytest.fixture
def db_manager():
    """Create a mock DatabaseManager instance."""
    return MockDatabaseManager()


@pytest.fixture
def pipeline(db_manager):
    """Create an ORSAPipeline instance."""
    return ORSAPipeline(db_manager, force_reprocess=False)


class TestORSAPipeline:
    """Test cases for ORSAPipeline class."""
    
    def test_initialization(self, pipeline, db_manager):
        """Test pipeline initialization."""
        assert pipeline.db_manager == db_manager
        assert pipeline.processor is not None
        assert pipeline.processing_stats["files_processed"] == 0
        assert pipeline.processing_stats["files_skipped"] == 0
        assert pipeline.processing_stats["files_failed"] == 0
    
    def test_initialization_with_force_reprocess(self, db_manager):
        """Test initialization with force reprocess mode."""
        pipeline = ORSAPipeline(db_manager, force_reprocess=True)
        assert pipeline.processor.force_reprocess is True
    
    def test_process_single_document(self, pipeline, sample_excel_file):
        """Test processing a single document."""
        documents = [("INST001_report.xlsx", sample_excel_file, None)]
        summary = pipeline.process_documents(documents)
        
        assert summary["files_processed"] == 1
        assert summary["files_skipped"] == 0
        assert summary["files_failed"] == 0
        assert summary["total_checks"] > 0
        assert summary["checks_passed"] >= 0
        assert summary["checks_failed"] >= 0
        assert summary["checks_passed"] + summary["checks_failed"] == summary["total_checks"]
        assert 0.0 <= summary["pass_rate"] <= 1.0
        assert "INST001" in summary["institutes"]
        assert "processing_time" in summary
        assert summary["processing_time"] >= 0
    
    def test_process_multiple_documents(
        self, pipeline, sample_excel_file, sample_excel_file2
    ):
        """Test processing multiple documents."""
        documents = [
            ("INST001_report.xlsx", sample_excel_file, None),
            ("INST002_report.xlsx", sample_excel_file2, None),
        ]
        summary = pipeline.process_documents(documents)
        
        assert summary["files_processed"] == 2
        assert summary["files_skipped"] == 0
        assert summary["files_failed"] == 0
        assert "INST001" in summary["institutes"]
        assert "INST002" in summary["institutes"]
    
    def test_process_duplicate_document_skipped(self, db_manager, sample_excel_file):
        """Test that duplicate documents are skipped."""
        # Create new pipeline for this test
        pipeline = ORSAPipeline(db_manager, force_reprocess=False)
        documents = [("INST001_report.xlsx", sample_excel_file, None)]
        
        # First processing
        summary1 = pipeline.process_documents(documents)
        assert summary1["files_processed"] == 1
        
        # Create another pipeline instance with same db_manager
        pipeline2 = ORSAPipeline(db_manager, force_reprocess=False)
        # Load existing versions
        pipeline2.processor.version_manager.load_existing_versions(
            db_manager.get_existing_versions()
        )
        
        # Second processing - should skip
        summary2 = pipeline2.process_documents(documents)
        assert summary2["files_processed"] == 0
        assert summary2["files_skipped"] == 1
    
    def test_process_duplicate_with_force_reprocess(
        self, db_manager, sample_excel_file
    ):
        """Test that force reprocess overrides caching."""
        # First pipeline without force
        pipeline1 = ORSAPipeline(db_manager, force_reprocess=False)
        documents = [("INST001_report.xlsx", sample_excel_file, None)]
        
        # First processing
        summary1 = pipeline1.process_documents(documents)
        assert summary1["files_processed"] == 1
        
        # Second pipeline with force reprocess
        pipeline2 = ORSAPipeline(db_manager, force_reprocess=True)
        pipeline2.processor.version_manager.load_existing_versions(
            db_manager.get_existing_versions()
        )
        
        # Second processing - should NOT skip
        summary2 = pipeline2.process_documents(documents)
        assert summary2["files_processed"] == 1
        assert summary2["files_skipped"] == 0
    
    def test_process_nonexistent_file(self, pipeline, tmp_path):
        """Test handling of non-existent file."""
        nonexistent = tmp_path / "nonexistent.xlsx"
        documents = [("INST001_nonexistent.xlsx", nonexistent, None)]
        
        summary = pipeline.process_documents(documents)
        
        assert summary["files_processed"] == 0
        assert summary["files_failed"] == 1
    
    def test_process_from_sourcer(self, pipeline, sample_excel_file):
        """Test processing documents from a sourcer."""
        documents = [("INST001_report.xlsx", sample_excel_file, None)]
        sourcer = MockORSADocumentSourcer(documents)
        
        summary = pipeline.process_from_sourcer(sourcer)
        
        assert summary["files_processed"] == 1
        assert "INST001" in summary["institutes"]
    
    def test_generate_summary(self, pipeline, sample_excel_file):
        """Test generating pipeline summary."""
        documents = [("INST001_report.xlsx", sample_excel_file, None)]
        pipeline.process_documents(documents)
        
        summary = pipeline.generate_summary()
        
        assert "total_files" in summary
        assert "institutes" in summary
        assert "pipeline_stats" in summary
        assert summary["pipeline_stats"]["files_processed"] == 1
    
    def test_generate_summary_empty(self, pipeline):
        """Test generating summary with no processing."""
        summary = pipeline.generate_summary()
        
        assert "total_files" in summary
        assert summary["total_files"] == 0
    
    def test_close(self, pipeline):
        """Test closing the pipeline."""
        # Should not raise any exceptions
        pipeline.close()
    
    def test_institute_id_extraction(self, pipeline, tmp_path):
        """Test extraction of institute IDs from filenames."""
        # Test different filename formats
        test_cases = [
            ("INST001_report.xlsx", "INST001"),
            ("INST002-report.xlsx", "INST002"),
            ("INST003 report.xlsx", "INST003"),
        ]
        
        for filename, expected_id in test_cases:
            file_path = tmp_path / filename
            wb = Workbook()
            wb.save(file_path)
            
            documents = [(filename, file_path, None)]
            summary = pipeline.process_documents(documents)
            
            assert expected_id in summary["institutes"]
    
    def test_empty_document_list(self, pipeline):
        """Test processing empty document list."""
        summary = pipeline.process_documents([])
        
        assert summary["files_processed"] == 0
        assert summary["files_skipped"] == 0
        assert summary["files_failed"] == 0
        assert len(summary["institutes"]) == 0


class TestCachedDocumentProcessor:
    """Test cases for CachedDocumentProcessor class."""
    
    def test_initialization(self, db_manager):
        """Test cached processor initialization."""
        processor = CachedDocumentProcessor(db_manager, cache_enabled=True)
        assert processor.db_manager == db_manager
        assert processor.cache_enabled is True
        assert processor.processor is not None
    
    def test_initialization_cache_disabled(self, db_manager):
        """Test initialization with cache disabled."""
        processor = CachedDocumentProcessor(db_manager, cache_enabled=False)
        assert processor.cache_enabled is False
        assert processor.processor.force_reprocess is True
    
    def test_get_cache_status_uncached(self, db_manager, sample_excel_file):
        """Test getting cache status for uncached file."""
        processor = CachedDocumentProcessor(db_manager)
        status = processor.get_cache_status("INST001", sample_excel_file)
        
        assert "is_cached" in status
        assert status["is_cached"] is False
        assert "file_hash" in status
        assert len(status["file_hash"]) == 64  # SHA-256 hash length
    
    def test_get_cache_status_cached(self, db_manager, sample_excel_file):
        """Test getting cache status for cached file."""
        processor = CachedDocumentProcessor(db_manager)
        
        # Process file to cache it
        processor.processor.process_file("INST001", sample_excel_file)
        
        # Check cache status
        status = processor.get_cache_status("INST001", sample_excel_file)
        
        assert status["is_cached"] is True
        assert "version_number" in status
        assert status["version_number"] == 1
    
    def test_invalidate_cache_specific_institute(
        self, db_manager, sample_excel_file, sample_excel_file2
    ):
        """Test invalidating cache for a specific institute."""
        processor = CachedDocumentProcessor(db_manager)
        
        # Process files for two institutes
        processor.processor.process_file("INST001", sample_excel_file)
        processor.processor.process_file("INST002", sample_excel_file2)
        
        # Invalidate cache for INST001
        processor.invalidate_cache("INST001")
        
        # INST001 should not be cached, INST002 should still be cached
        status1 = processor.get_cache_status("INST001", sample_excel_file)
        status2 = processor.get_cache_status("INST002", sample_excel_file2)
        
        assert status1["is_cached"] is False
        assert status2["is_cached"] is True
    
    def test_invalidate_cache_all(
        self, db_manager, sample_excel_file, sample_excel_file2
    ):
        """Test invalidating entire cache."""
        processor = CachedDocumentProcessor(db_manager)
        
        # Process files
        processor.processor.process_file("INST001", sample_excel_file)
        processor.processor.process_file("INST002", sample_excel_file2)
        
        # Invalidate all cache
        processor.invalidate_cache()
        
        # Both should not be cached
        status1 = processor.get_cache_status("INST001", sample_excel_file)
        status2 = processor.get_cache_status("INST002", sample_excel_file2)
        
        assert status1["is_cached"] is False
        assert status2["is_cached"] is False
    
    def test_get_cache_statistics_empty(self, db_manager):
        """Test cache statistics with empty cache."""
        processor = CachedDocumentProcessor(db_manager)
        stats = processor.get_cache_statistics()
        
        assert stats["total_institutes"] == 0
        assert stats["total_versions"] == 0
        assert len(stats["institutes"]) == 0
    
    def test_get_cache_statistics_with_data(
        self, db_manager, sample_excel_file, sample_excel_file2
    ):
        """Test cache statistics with cached data."""
        processor = CachedDocumentProcessor(db_manager)
        
        # Process files
        processor.processor.process_file("INST001", sample_excel_file)
        processor.processor.process_file("INST002", sample_excel_file2)
        
        stats = processor.get_cache_statistics()
        
        assert stats["total_institutes"] == 2
        assert stats["total_versions"] == 2
        assert "INST001" in stats["institutes"]
        assert "INST002" in stats["institutes"]
    
    def test_cache_statistics_after_invalidation(self, db_manager, sample_excel_file):
        """Test cache statistics after cache invalidation."""
        processor = CachedDocumentProcessor(db_manager)
        
        # Process and cache
        processor.processor.process_file("INST001", sample_excel_file)
        stats1 = processor.get_cache_statistics()
        assert stats1["total_versions"] == 1
        
        # Invalidate and check
        processor.invalidate_cache()
        stats2 = processor.get_cache_statistics()
        assert stats2["total_versions"] == 0
