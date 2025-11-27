"""Tests for TemplateManager class."""

import pytest
from pathlib import Path
import openpyxl
from openpyxl.workbook.workbook import Workbook

from orsa_analysis.reporting.template_manager import TemplateManager


@pytest.fixture
def temp_template_file(tmp_path):
    """Create a temporary template Excel file."""
    template_path = tmp_path / "test_template.xlsx"
    wb = Workbook()
    
    # Remove default sheet and add Auswertung sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    ws = wb.create_sheet("Auswertung")
    ws["A1"] = "Header"
    ws["C8"] = "Original Value"
    
    wb.save(template_path)
    return template_path


@pytest.fixture
def temp_source_file(tmp_path):
    """Create a temporary source Excel file."""
    source_path = tmp_path / "test_source.xlsx"
    wb = Workbook()
    
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Add multiple sheets
    ws1 = wb.create_sheet("Data Sheet 1")
    ws1["A1"] = "Data 1"
    
    ws2 = wb.create_sheet("Data Sheet 2")
    ws2["A1"] = "Data 2"
    
    wb.save(source_path)
    return source_path


class TestTemplateManagerInitialization:
    """Test TemplateManager initialization."""
    
    def test_init_with_valid_template(self, temp_template_file):
        """Test initialization with valid template file."""
        manager = TemplateManager(temp_template_file)
        assert manager.template_path == temp_template_file
    
    def test_init_with_nonexistent_template(self, tmp_path):
        """Test initialization with non-existent template raises error."""
        nonexistent = tmp_path / "nonexistent.xlsx"
        with pytest.raises(FileNotFoundError):
            TemplateManager(nonexistent)


class TestTemplateLoading:
    """Test template and source file loading."""
    
    def test_load_template(self, temp_template_file):
        """Test loading template file."""
        manager = TemplateManager(temp_template_file)
        wb = manager.load_template()
        
        assert isinstance(wb, Workbook)
        assert "Auswertung" in wb.sheetnames
        assert wb["Auswertung"]["A1"].value == "Header"
    
    def test_load_source_file_exists(self, temp_template_file, temp_source_file):
        """Test loading existing source file."""
        manager = TemplateManager(temp_template_file)
        wb = manager.load_source_file(temp_source_file)
        
        assert wb is not None
        assert isinstance(wb, Workbook)
        assert "Data Sheet 1" in wb.sheetnames
    
    def test_load_source_file_not_exists(self, temp_template_file, tmp_path):
        """Test loading non-existent source file returns None."""
        manager = TemplateManager(temp_template_file)
        nonexistent = tmp_path / "nonexistent.xlsx"
        wb = manager.load_source_file(nonexistent)
        
        assert wb is None


class TestOutputWorkbookCreation:
    """Test output workbook creation."""
    
    def test_create_output_template_only(self, temp_template_file):
        """Test creating output workbook from template only."""
        manager = TemplateManager(temp_template_file)
        template_wb = manager.load_template()
        
        output_wb = manager.create_output_workbook(template_wb)
        
        assert isinstance(output_wb, Workbook)
        assert "Auswertung" in output_wb.sheetnames
        assert output_wb["Auswertung"]["A1"].value == "Header"
    
    def test_create_output_with_source(self, temp_template_file, temp_source_file):
        """Test creating output workbook with source sheets."""
        manager = TemplateManager(temp_template_file)
        template_wb = manager.load_template()
        source_wb = manager.load_source_file(temp_source_file)
        
        output_wb = manager.create_output_workbook(template_wb, source_wb)
        
        # Should have template sheets first
        assert "Auswertung" in output_wb.sheetnames
        # Then source sheets
        assert "Data Sheet 1" in output_wb.sheetnames
        assert "Data Sheet 2" in output_wb.sheetnames
        
        # Template sheets should come first
        assert output_wb.sheetnames.index("Auswertung") < output_wb.sheetnames.index("Data Sheet 1")
    
    def test_create_output_sheet_name_conflict(self, temp_template_file, tmp_path):
        """Test handling of sheet name conflicts."""
        # Create source file with same sheet name as template
        source_path = tmp_path / "conflict_source.xlsx"
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        ws = wb.create_sheet("Auswertung")
        ws["A1"] = "Conflicting Sheet"
        wb.save(source_path)
        
        manager = TemplateManager(temp_template_file)
        template_wb = manager.load_template()
        source_wb = manager.load_source_file(source_path)
        
        output_wb = manager.create_output_workbook(template_wb, source_wb)
        
        # Should have both sheets with different names
        assert "Auswertung" in output_wb.sheetnames
        assert "Auswertung_Source" in output_wb.sheetnames


class TestCellOperations:
    """Test writing to cells."""
    
    def test_write_cell_value_success(self, temp_template_file):
        """Test writing value to existing sheet."""
        manager = TemplateManager(temp_template_file)
        wb = manager.load_template()
        
        success = manager.write_cell_value(wb, "Auswertung", "C8", "New Value")
        
        assert success is True
        assert wb["Auswertung"]["C8"].value == "New Value"
    
    def test_write_cell_value_nonexistent_sheet(self, temp_template_file):
        """Test writing to non-existent sheet returns False."""
        manager = TemplateManager(temp_template_file)
        wb = manager.load_template()
        
        success = manager.write_cell_value(wb, "NonExistent", "A1", "Value")
        
        assert success is False
    
    def test_write_cell_value_various_types(self, temp_template_file):
        """Test writing various value types."""
        manager = TemplateManager(temp_template_file)
        wb = manager.load_template()
        ws = wb["Auswertung"]
        
        # String
        manager.write_cell_value(wb, "Auswertung", "A2", "Text")
        assert ws["A2"].value == "Text"
        
        # Number
        manager.write_cell_value(wb, "Auswertung", "A3", 123.45)
        assert ws["A3"].value == 123.45
        
        # Boolean
        manager.write_cell_value(wb, "Auswertung", "A4", True)
        assert ws["A4"].value is True
        
        # None
        manager.write_cell_value(wb, "Auswertung", "A5", None)
        assert ws["A5"].value is None


class TestWorkbookSaving:
    """Test saving workbooks."""
    
    def test_save_workbook_success(self, temp_template_file, tmp_path):
        """Test successfully saving workbook."""
        manager = TemplateManager(temp_template_file)
        wb = manager.load_template()
        output_path = tmp_path / "output" / "test_output.xlsx"
        
        success = manager.save_workbook(wb, output_path)
        
        assert success is True
        assert output_path.exists()
        
        # Verify file can be opened
        saved_wb = openpyxl.load_workbook(output_path)
        assert "Auswertung" in saved_wb.sheetnames
    
    def test_save_workbook_creates_directory(self, temp_template_file, tmp_path):
        """Test that save creates output directory if it doesn't exist."""
        manager = TemplateManager(temp_template_file)
        wb = manager.load_template()
        output_path = tmp_path / "new_dir" / "subdir" / "output.xlsx"
        
        success = manager.save_workbook(wb, output_path)
        
        assert success is True
        assert output_path.parent.exists()
        assert output_path.exists()


class TestWorkbookClosing:
    """Test closing workbooks."""
    
    def test_close_workbook(self, temp_template_file):
        """Test closing workbook."""
        manager = TemplateManager(temp_template_file)
        wb = manager.load_template()
        
        # Should not raise error
        manager.close_workbook(wb)


class TestCopySheet:
    """Test internal _copy_sheet method."""
    
    def test_copy_sheet_preserves_values(self, temp_template_file):
        """Test that copying preserves cell values."""
        manager = TemplateManager(temp_template_file)
        template_wb = manager.load_template()
        
        # Add some values to template
        source_sheet = template_wb["Auswertung"]
        source_sheet["B2"] = "Test Value"
        source_sheet["C3"] = 42
        
        # Create new workbook and copy sheet
        target_wb = Workbook()
        if "Sheet" in target_wb.sheetnames:
            target_wb.remove(target_wb["Sheet"])
        
        copied_sheet = manager._copy_sheet(source_sheet, target_wb, "Copied")
        
        assert copied_sheet["B2"].value == "Test Value"
        assert copied_sheet["C3"].value == 42
