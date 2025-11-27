"""Unit tests for the processor module."""

import pytest
from pathlib import Path
from openpyxl import Workbook
from typing import List, Dict, Any

from orsa_analysis.core.processor import DocumentProcessor
from orsa_analysis.core.database_manager import CheckResult


class MockDatabaseManager:
    """Mock DatabaseManager for testing."""
    
    def __init__(self):
        self.stored_results = []
        self.existing_versions = []
    
    def write_results(self, results: List[CheckResult]) -> int:
        self.stored_results.extend(results)
        return len(results)
    
    def get_existing_versions(self) -> List[Dict[str, Any]]:
        return self.existing_versions.copy()
    
    def close(self):
        pass


@pytest.fixture
def sample_excel_file(tmp_path):
    """Create a sample Excel file for testing."""
    file_path = tmp_path / "INST001_report.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    ws["A1"] = "Header1"
    ws["B1"] = "Header2"
    ws["A2"] = "Data1"
    ws["B2"] = "Data2"
    wb.save(file_path)
    return file_path


@pytest.fixture
def db_manager():
    """Create a mock DatabaseManager instance."""
    return MockDatabaseManager()


@pytest.fixture
def processor(db_manager):
    """Create a DocumentProcessor instance."""
    return DocumentProcessor(db_manager, force_reprocess=False)


class TestDocumentProcessor:
    """Test cases for DocumentProcessor class."""

    def test_initialization(self, processor):
        """Test DocumentProcessor initialization."""
        assert processor.reader is not None
        assert processor.version_manager is not None
        assert processor.db_manager is not None
        assert processor.force_reprocess is False

    def test_initialization_with_force(self, db_manager):
        """Test initialization with force reprocess."""
        processor = DocumentProcessor(db_manager, force_reprocess=True)
        assert processor.force_reprocess is True

    def test_should_process_file_new(self, processor, sample_excel_file):
        """Test should_process_file for a new file."""
        should_process, reason = processor.should_process_file(
            "INST001", sample_excel_file
        )

        assert should_process is True
        assert "new" in reason.lower() or "hash" in reason.lower()

    def test_should_process_file_force_mode(self, db_manager, sample_excel_file):
        """Test should_process_file in force mode."""
        processor = DocumentProcessor(db_manager, force_reprocess=True)

        should_process, reason = processor.should_process_file(
            "INST001", sample_excel_file
        )

        assert should_process is True
        assert "force" in reason.lower()

    def test_should_process_file_already_processed(
        self, processor, sample_excel_file
    ):
        """Test should_process_file for an already processed file."""
        processor.process_file("INST001", sample_excel_file)

        should_process, reason = processor.should_process_file(
            "INST001", sample_excel_file
        )

        assert should_process is False
        assert "already processed" in reason.lower()

    def test_process_file(self, processor, sample_excel_file):
        """Test processing a single file."""
        version_info, results = processor.process_file("INST001", sample_excel_file)

        assert version_info.institute_id == "INST001"
        assert version_info.version_number == 1
        assert len(results) > 0

        for result in results:
            assert result.institute_id == "INST001"
            assert result.file_name == sample_excel_file.name
            assert result.file_hash == version_info.file_hash

    def test_process_file_multiple_checks(self, processor, sample_excel_file):
        """Test that multiple checks are executed."""
        _, results = processor.process_file("INST001", sample_excel_file)

        check_names = [r.check_name for r in results]
        assert len(check_names) > 5
        assert "has_sheets" in check_names
        assert "no_empty_sheets" in check_names

    def test_process_file_not_found(self, processor):
        """Test processing a non-existent file."""
        with pytest.raises(FileNotFoundError):
            processor.process_file("INST001", Path("/nonexistent/file.xlsx"))

    def test_process_documents(self, processor, tmp_path):
        """Test processing multiple documents."""
        file1 = tmp_path / "INST001_report.xlsx"
        file2 = tmp_path / "INST002_report.xlsx"

        for file_path in [file1, file2]:
            wb = Workbook()
            wb.active["A1"] = "Data"
            wb.save(file_path)

        documents = [(file1.name, file1), (file2.name, file2)]

        results = processor.process_documents(documents)

        assert len(results) == 2
        assert results[0][0] == "INST001"
        assert results[1][0] == "INST002"

    def test_process_documents_with_skip(self, processor, tmp_path):
        """Test processing documents with some being skipped."""
        file1 = tmp_path / "INST001_report.xlsx"

        wb = Workbook()
        wb.active["A1"] = "Data"
        wb.save(file1)

        documents = [(file1.name, file1)]

        results1 = processor.process_documents(documents)
        assert len(results1) == 1

        results2 = processor.process_documents(documents)
        assert len(results2) == 0

    def test_extract_institute_id_underscore(self, processor):
        """Test extracting institute ID with underscore separator."""
        institute_id = processor._extract_institute_id("INST001_report.xlsx")
        assert institute_id == "INST001"

    def test_extract_institute_id_dash(self, processor):
        """Test extracting institute ID with dash separator."""
        institute_id = processor._extract_institute_id("INST002-report.xlsx")
        assert institute_id == "INST002"

    def test_extract_institute_id_space(self, processor):
        """Test extracting institute ID with space separator."""
        institute_id = processor._extract_institute_id("INST003 report.xlsx")
        assert institute_id == "INST003"

    def test_extract_institute_id_no_separator(self, processor):
        """Test extracting institute ID without separator."""
        institute_id = processor._extract_institute_id("INST004.xlsx")
        assert institute_id == "INST004"

    def test_get_processing_summary_empty(self, processor):
        """Test getting summary with no processed files."""
        summary = processor.get_processing_summary()

        assert summary["total_files"] == 0
        assert summary["total_checks"] == 0
        assert summary["checks_passed"] == 0
        assert summary["checks_failed"] == 0
        assert summary["institutes"] == []

    def test_get_processing_summary_with_data(self, processor, sample_excel_file):
        """Test getting summary after processing files."""
        processor.process_file("INST001", sample_excel_file)

        summary = processor.get_processing_summary()

        assert summary["total_files"] >= 1
        # Note: In the simplified version, total_checks is not tracked
        assert summary["total_checks"] == 0
        assert "INST001" in summary["institutes"]
        assert summary["pass_rate"] == "N/A"

    def test_versioning_increments(self, processor, tmp_path):
        """Test that version numbers increment for new files."""
        file1 = tmp_path / "INST001_v1.xlsx"
        file2 = tmp_path / "INST001_v2.xlsx"

        wb1 = Workbook()
        wb1.active["A1"] = "Data1"
        wb1.save(file1)

        wb2 = Workbook()
        wb2.active["A1"] = "Data2"
        wb2.save(file2)

        version1, _ = processor.process_file("INST001", file1)
        version2, _ = processor.process_file("INST001", file2)

        assert version1.version_number == 1
        assert version2.version_number == 2

    def test_versioning_same_hash(self, processor, sample_excel_file):
        """Test that same file gets same version number."""
        version1, _ = processor.process_file("INST001", sample_excel_file)

        mock_db = MockDatabaseManager()
        processor_force = DocumentProcessor(mock_db, force_reprocess=True)
        processor_force.version_manager.load_existing_versions(
            processor.db_manager.get_existing_versions()
        )

        version2, _ = processor_force.process_file("INST001", sample_excel_file)

        assert version1.version_number == version2.version_number
        assert version1.file_hash == version2.file_hash

    def test_results_written_to_db(self, processor, sample_excel_file):
        """Test that results are written to database."""
        processor.process_file("INST001", sample_excel_file)

        stored_results = processor.db_manager.stored_results
        assert len(stored_results) > 0

    def test_different_institutes_independent_versions(self, processor, tmp_path):
        """Test that different institutes have independent version numbers."""
        file1 = tmp_path / "INST001_report.xlsx"
        file2 = tmp_path / "INST002_report.xlsx"

        wb1 = Workbook()
        wb1.active["A1"] = "Data"
        wb1.save(file1)

        wb2 = Workbook()
        wb2.active["A1"] = "Data"
        wb2.save(file2)

        version1, _ = processor.process_file("INST001", file1)
        version2, _ = processor.process_file("INST002", file2)

        assert version1.version_number == 1
        assert version2.version_number == 1
