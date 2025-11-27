"""Unit tests for the reader module."""

import pytest
from pathlib import Path
from openpyxl import Workbook

from orsa_analysis.core.reader import ExcelReader


@pytest.fixture
def sample_excel_file(tmp_path):
    """Create a sample Excel file for testing."""
    file_path = tmp_path / "test_file.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    ws["A1"] = "Header1"
    ws["B1"] = "Header2"
    ws["A2"] = "Data1"
    ws["B2"] = "Data2"
    wb.save(file_path)
    return file_path


@pytest.fixture
def empty_excel_file(tmp_path):
    """Create an empty Excel file for testing."""
    file_path = tmp_path / "empty_file.xlsx"
    wb = Workbook()
    wb.save(file_path)
    return file_path


class TestExcelReader:
    """Test cases for ExcelReader class."""

    def test_initialization_default(self):
        """Test ExcelReader initialization with default parameters."""
        reader = ExcelReader()
        assert reader.data_only is True
        assert reader.read_only is False

    def test_initialization_custom(self):
        """Test ExcelReader initialization with custom parameters."""
        reader = ExcelReader(data_only=False, read_only=True)
        assert reader.data_only is False
        assert reader.read_only is True

    def test_load_file_success(self, sample_excel_file):
        """Test successful loading of an Excel file."""
        reader = ExcelReader()
        workbook = reader.load_file(sample_excel_file)

        assert workbook is not None
        assert "TestSheet" in workbook.sheetnames
        reader.close_workbook(workbook)

    def test_load_file_not_found(self):
        """Test loading a non-existent file."""
        reader = ExcelReader()
        with pytest.raises(FileNotFoundError):
            reader.load_file(Path("/nonexistent/file.xlsx"))

    def test_load_file_invalid_extension(self, tmp_path):
        """Test loading a file with invalid extension."""
        invalid_file = tmp_path / "test.txt"
        invalid_file.write_text("Not an Excel file")

        reader = ExcelReader()
        with pytest.raises(ValueError, match="Invalid Excel file extension"):
            reader.load_file(invalid_file)

    def test_get_sheet_names(self, sample_excel_file):
        """Test retrieving sheet names from a workbook."""
        reader = ExcelReader()
        workbook = reader.load_file(sample_excel_file)

        sheet_names = reader.get_sheet_names(workbook)
        assert isinstance(sheet_names, list)
        assert len(sheet_names) >= 1
        assert "TestSheet" in sheet_names

        reader.close_workbook(workbook)

    def test_close_workbook(self, sample_excel_file):
        """Test closing a workbook."""
        reader = ExcelReader()
        workbook = reader.load_file(sample_excel_file)

        reader.close_workbook(workbook)

    def test_load_empty_file(self, empty_excel_file):
        """Test loading an empty Excel file."""
        reader = ExcelReader()
        workbook = reader.load_file(empty_excel_file)

        assert workbook is not None
        assert len(workbook.sheetnames) >= 1

        reader.close_workbook(workbook)

    def test_multiple_sheets(self, tmp_path):
        """Test loading a file with multiple sheets."""
        file_path = tmp_path / "multi_sheet.xlsx"
        wb = Workbook()
        wb.create_sheet("Sheet2")
        wb.create_sheet("Sheet3")
        wb.save(file_path)

        reader = ExcelReader()
        workbook = reader.load_file(file_path)

        sheet_names = reader.get_sheet_names(workbook)
        assert len(sheet_names) == 3

        reader.close_workbook(workbook)
