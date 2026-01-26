"""Tests for ExcelTemplateManager class."""

import pytest
from pathlib import Path
import openpyxl
from openpyxl.workbook.workbook import Workbook

from orsa_analysis.reporting.excel_template_manager import ExcelTemplateManager


@pytest.fixture
def temp_template_file(tmp_path):
    """Create a temporary template Excel file."""
    template_path = tmp_path / "test_template.xlsx"
    wb = Workbook()
    
    # Remove default sheet and add Auswertung sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    ws = wb.create_sheet("Auswertung")
    ws["A1"] = "Header"
    ws["C8"] = "Original Value"
    
    wb.save(template_path)
    wb.close()
    return template_path


@pytest.fixture
def temp_source_file(tmp_path):
    """Create a temporary source Excel file."""
    source_path = tmp_path / "test_source.xlsx"
    wb = Workbook()
    
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Add multiple sheets
    ws1 = wb.create_sheet("Data Sheet 1")
    ws1["A1"] = "Data 1"
    
    ws2 = wb.create_sheet("Data Sheet 2")
    ws2["A1"] = "Data 2"
    
    wb.save(source_path)
    wb.close()
    return source_path


class TestExcelTemplateManagerInitialization:
    """Test ExcelTemplateManager initialization."""
    
    def test_init_with_valid_template(self, temp_template_file):
        """Test initialization with valid template file."""
        manager = ExcelTemplateManager(temp_template_file)
        assert manager.template_path == temp_template_file
    
    def test_init_with_nonexistent_template(self, tmp_path):
        """Test initialization with non-existent template raises error."""
        nonexistent = tmp_path / "nonexistent.xlsx"
        with pytest.raises(FileNotFoundError):
            ExcelTemplateManager(nonexistent)


class TestOutputWorkbookCreation:
    """Test output workbook creation."""
    
    def test_create_output_workbook(self, temp_template_file, temp_source_file):
        """Test creating standalone output workbook from template only."""
        manager = ExcelTemplateManager(temp_template_file)
        wb = manager.create_output_workbook(temp_source_file)
        
        assert isinstance(wb, Workbook)
        # Should only contain template sheet
        assert "Auswertung" in wb.sheetnames
        assert wb.sheetnames[0] == "Auswertung"
        
        # Source sheets should NOT be included
        assert "Data Sheet 1" not in wb.sheetnames
        assert "Data Sheet 2" not in wb.sheetnames
        
        # Total should be only template sheets
        assert len(wb.sheetnames) == 1
    
    def test_create_output_workbook_nonexistent_source(self, temp_template_file, tmp_path):
        """Test creating output workbook with non-existent source file raises error."""
        manager = ExcelTemplateManager(temp_template_file)
        nonexistent = tmp_path / "nonexistent.xlsx"
        
        with pytest.raises(FileNotFoundError):
            manager.create_output_workbook(nonexistent)
    
    def test_create_output_preserves_template_content(self, temp_template_file, temp_source_file):
        """Test that template content is preserved in output."""
        manager = ExcelTemplateManager(temp_template_file)
        wb = manager.create_output_workbook(temp_source_file)
        
        # Check template content is preserved
        assert wb["Auswertung"]["A1"].value == "Header"
        assert wb["Auswertung"]["C8"].value == "Original Value"
    
    def test_create_output_does_not_include_source_content(self, temp_template_file, temp_source_file):
        """Test that source content is NOT included in output (standalone report)."""
        manager = ExcelTemplateManager(temp_template_file)
        wb = manager.create_output_workbook(temp_source_file)
        
        # Check that source sheets are NOT present
        assert "Data Sheet 1" not in wb.sheetnames
        assert "Data Sheet 2" not in wb.sheetnames


class TestCellOperations:
    """Test cell write operations."""
    
    def test_write_cell_value_success(self, temp_template_file, temp_source_file):
        """Test writing value to cell."""
        manager = ExcelTemplateManager(temp_template_file)
        manager.create_output_workbook(temp_source_file)
        
        success = manager.write_cell_value("Auswertung", "C8", "New Value")
        assert success is True
        assert manager.output_wb["Auswertung"]["C8"].value == "New Value"
    
    def test_write_cell_value_nonexistent_sheet(self, temp_template_file, temp_source_file):
        """Test writing to non-existent sheet returns False."""
        manager = ExcelTemplateManager(temp_template_file)
        manager.create_output_workbook(temp_source_file)
        
        success = manager.write_cell_value("NonExistent", "A1", "Value")
        assert success is False
    
    def test_write_cell_value_no_workbook(self, temp_template_file):
        """Test writing without creating workbook returns False."""
        manager = ExcelTemplateManager(temp_template_file)
        success = manager.write_cell_value("Auswertung", "A1", "Value")
        assert success is False
    
    def test_write_cell_value_various_types(self, temp_template_file, temp_source_file):
        """Test writing various data types to cells."""
        manager = ExcelTemplateManager(temp_template_file)
        manager.create_output_workbook(temp_source_file)
        
        # Test different value types
        test_cases = [
            ("A1", "Pass", str),
            ("A2", 123, int),
            ("A3", 45.67, float),
            ("A4", True, bool),
        ]
        
        for cell, value, expected_type in test_cases:
            success = manager.write_cell_value("Auswertung", cell, value)
            assert success is True
            cell_value = manager.output_wb["Auswertung"][cell].value
            assert isinstance(cell_value, expected_type)
            assert cell_value == value
    
    def test_write_cell_value_numeric_string_to_int(self, temp_template_file, temp_source_file):
        """Test that numeric strings are converted to integers."""
        manager = ExcelTemplateManager(temp_template_file)
        manager.create_output_workbook(temp_source_file)
        
        # Write a string representation of an integer
        success = manager.write_cell_value("Auswertung", "B1", "42")
        assert success is True
        
        # Verify it was converted to an actual integer
        cell_value = manager.output_wb["Auswertung"]["B1"].value
        assert isinstance(cell_value, int)
        assert cell_value == 42
    
    def test_write_cell_value_numeric_string_to_float(self, temp_template_file, temp_source_file):
        """Test that numeric strings with decimals are converted to floats."""
        manager = ExcelTemplateManager(temp_template_file)
        manager.create_output_workbook(temp_source_file)
        
        # Write a string representation of a float
        success = manager.write_cell_value("Auswertung", "B2", "42.5")
        assert success is True
        
        # Verify it was converted to an actual float
        cell_value = manager.output_wb["Auswertung"]["B2"].value
        assert isinstance(cell_value, float)
        assert cell_value == 42.5
    
    def test_write_cell_value_negative_numeric_strings(self, temp_template_file, temp_source_file):
        """Test that negative numeric strings are converted correctly."""
        manager = ExcelTemplateManager(temp_template_file)
        manager.create_output_workbook(temp_source_file)
        
        # Test negative integer
        success = manager.write_cell_value("Auswertung", "B3", "-15")
        assert success is True
        cell_value = manager.output_wb["Auswertung"]["B3"].value
        assert isinstance(cell_value, int)
        assert cell_value == -15
        
        # Test negative float
        success = manager.write_cell_value("Auswertung", "B4", "-15.75")
        assert success is True
        cell_value = manager.output_wb["Auswertung"]["B4"].value
        assert isinstance(cell_value, float)
        assert cell_value == -15.75
    
    def test_write_cell_value_non_numeric_strings_unchanged(self, temp_template_file, temp_source_file):
        """Test that non-numeric strings remain as strings."""
        manager = ExcelTemplateManager(temp_template_file)
        manager.create_output_workbook(temp_source_file)
        
        # Write non-numeric strings
        test_cases = [
            ("B5", "genügend"),
            ("B6", "zu prüfen"),
            ("B7", "abc123"),
            ("B8", "12.34.56"),  # Invalid number format
        ]
        
        for cell, value in test_cases:
            success = manager.write_cell_value("Auswertung", cell, value)
            assert success is True
            cell_value = manager.output_wb["Auswertung"][cell].value
            assert isinstance(cell_value, str)
            assert cell_value == value
    
    def test_write_cell_value_empty_string_unchanged(self, temp_template_file, temp_source_file):
        """Test that empty strings remain as empty strings."""
        manager = ExcelTemplateManager(temp_template_file)
        manager.create_output_workbook(temp_source_file)
        
        # Write empty string
        success = manager.write_cell_value("Auswertung", "B9", "")
        assert success is True
        cell_value = manager.output_wb["Auswertung"]["B9"].value
        # Empty string in openpyxl might be None or ""
        assert cell_value in ["", None]


class TestWorkbookSaving:
    """Test workbook save operations."""
    
    def test_save_workbook_success(self, temp_template_file, temp_source_file, tmp_path):
        """Test saving workbook to file."""
        manager = ExcelTemplateManager(temp_template_file)
        manager.create_output_workbook(temp_source_file)
        
        output_path = tmp_path / "output.xlsx"
        manager.save_workbook(output_path)
        
        assert output_path.exists()
        
        # Verify saved file is valid and contains only template sheets
        wb = openpyxl.load_workbook(output_path)
        assert "Auswertung" in wb.sheetnames
        assert "Data Sheet 1" not in wb.sheetnames
        wb.close()
    
    def test_save_workbook_creates_directory(self, temp_template_file, temp_source_file, tmp_path):
        """Test that save creates output directory if it doesn't exist."""
        manager = ExcelTemplateManager(temp_template_file)
        manager.create_output_workbook(temp_source_file)
        
        output_path = tmp_path / "subdir" / "output.xlsx"
        manager.save_workbook(output_path)
        
        assert output_path.exists()
        assert output_path.parent.exists()


class TestWorkbookClosing:
    """Test workbook close operations."""
    
    def test_close_workbook(self, temp_template_file, temp_source_file):
        """Test closing workbook."""
        manager = ExcelTemplateManager(temp_template_file)
        manager.create_output_workbook(temp_source_file)
        
        assert manager.output_wb is not None
        manager.close()
        assert manager.output_wb is None
