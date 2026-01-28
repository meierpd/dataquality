#!/usr/bin/env python3
"""Debug script to test a specific check on a specific file."""

from pathlib import Path
from openpyxl import load_workbook
from orsa_analysis.checks.rules import check_tied_assets_filled_three_years
from orsa_analysis.checks.sheet_mapper import SheetNameMapper

# ============================================
# CONFIGURATION - Update these values
# ============================================
FILE_PATH = "/path/to/your/file.xlsx"  # <-- CHANGE THIS
INSTITUTE_ID = "YOUR_INSTITUTE_ID"      # <-- CHANGE THIS (optional, for reference only)

def debug_check():
    """Run the check with detailed debugging output."""
    
    print(f"Loading file: {FILE_PATH}")
    print(f"Institute: {INSTITUTE_ID}")
    print("=" * 80)
    
    # Load the workbook
    wb = load_workbook(FILE_PATH, data_only=True, read_only=False)
    
    # Initialize mapper
    mapper = SheetNameMapper(wb)
    
    print("\n📋 AVAILABLE SHEETS:")
    for sheet_name in wb.sheetnames:
        print(f"  - {sheet_name}")
    
    # Check which sheet is being used
    print("\n🔍 DETERMINING SHEET TO USE:")
    
    # Check if Zweigniederlassungs version (handles all languages)
    zweigniederlassungs_sheets = {"Ergebnisse", "Results", "Résultats"}
    is_zweigniederlassungs = False
    
    for sheet_name in wb.sheetnames:
        if sheet_name in zweigniederlassungs_sheets:
            is_zweigniederlassungs = True
            result_sheet = wb[sheet_name]
            print(f"  ✓ Zweigniederlassungs version detected")
            print(f"  ✓ Using sheet: '{sheet_name}'")
            break
    
    if not is_zweigniederlassungs:
        # Standard version - check which sheet is filled
        sheet_avo = mapper.get_sheet("Ergebnisse_AVO-FINMA")
        sheet_ifrs = mapper.get_sheet("Ergebnisse_IFRS")
        
        # Check E26, F26, G26 to see which is filled
        check_cells = ["E26", "F26", "G26"]
        
        avo_filled = False
        if sheet_avo:
            avo_values = [sheet_avo[cell].value for cell in check_cells]
            avo_filled = any(v is not None and str(v).strip() != "" for v in avo_values)
            print(f"  AVO-FINMA sheet found, cells {check_cells}: {avo_values} - Filled: {avo_filled}")
        
        ifrs_filled = False
        if sheet_ifrs:
            ifrs_values = [sheet_ifrs[cell].value for cell in check_cells]
            ifrs_filled = any(v is not None and str(v).strip() != "" for v in ifrs_values)
            print(f"  IFRS sheet found, cells {check_cells}: {ifrs_values} - Filled: {ifrs_filled}")
        
        if avo_filled:
            result_sheet = sheet_avo
            print(f"  ✓ Using AVO-FINMA sheet")
            is_avo = True
        elif ifrs_filled:
            result_sheet = sheet_ifrs
            print(f"  ✓ Using IFRS sheet")
            is_avo = False
        else:
            print("  ✗ ERROR: No results sheet is filled!")
            return
    
    # Now check the tied assets cells
    print("\n📊 CHECKING TIED ASSETS (Gebundenes Vermögen):")
    
    if is_zweigniederlassungs:
        start_col, end_col = "E", "G"
        start_row, end_row = 38, 40
        print(f"  Range to check (Zweigniederlassungs): {start_col}{start_row}:{end_col}{end_row}")
    elif 'is_avo' in locals() and is_avo:
        start_col, end_col = "E", "G"
        start_row, end_row = 49, 51
        print(f"  Range to check (AVO-FINMA): {start_col}{start_row}:{end_col}{end_row}")
    else:
        start_col, end_col = "E", "G"
        start_row, end_row = 51, 54
        print(f"  Range to check (IFRS): {start_col}{start_row}:{end_col}{end_row}")
    
    # Check each cell
    empty_cells = []
    filled_cells = []
    
    for row in range(start_row, end_row + 1):
        for col_ord in range(ord(start_col), ord(end_col) + 1):
            cell_addr = f"{chr(col_ord)}{row}"
            cell_value = result_sheet[cell_addr].value
            is_empty = cell_value is None or str(cell_value).strip() == ""
            
            if is_empty:
                empty_cells.append(cell_addr)
                print(f"  ✗ {cell_addr}: EMPTY")
            else:
                filled_cells.append(cell_addr)
                print(f"  ✓ {cell_addr}: {cell_value}")
    
    # Run the actual check
    print("\n🎯 RUNNING ACTUAL CHECK:")
    outcome_bool, outcome_str, description = check_tied_assets_filled_three_years(wb)
    
    print(f"  Result: {'PASS ✓' if outcome_bool else 'FAIL ✗'}")
    print(f"  Outcome: {outcome_str}")
    print(f"  Description: {description}")
    
    # Summary
    print("\n📈 SUMMARY:")
    print(f"  Total cells checked: {len(filled_cells) + len(empty_cells)}")
    print(f"  Filled cells: {len(filled_cells)}")
    print(f"  Empty cells: {len(empty_cells)}")
    
    if empty_cells:
        print(f"\n  ⚠️  Empty cells found: {', '.join(empty_cells)}")
        print(f"  → Check needs ALL cells in range to be filled!")
    else:
        print(f"  ✓ All cells are filled")
    
    # Close workbook
    wb.close()
    print("\n" + "=" * 80)
    print("Done!")


if __name__ == "__main__":
    try:
        debug_check()
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
